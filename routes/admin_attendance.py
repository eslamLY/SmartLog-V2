import logging
from collections import defaultdict
from datetime import datetime, date
from io import BytesIO

from flask import (Blueprint, render_template, request, session,
                   jsonify, send_file, make_response)
from sqlalchemy import extract, or_

from models import db, Employee, AttendanceLog, Notification
from utils.decorators import admin_required
from utils.helpers import work_hours_str
from utils.constants import MONTH_NAMES
from services.payroll_service import PayrollService

logger = logging.getLogger(__name__)

admin_attendance_bp = Blueprint('admin_attendance', __name__)


@admin_attendance_bp.route('/admin/attendance')
@admin_required
def admin_attendance():
    today      = date.today()
    start_date = request.args.get('start_date', '')
    end_date   = request.args.get('end_date', '')
    sel_date   = request.args.get('date', '')
    sel_dept   = request.args.get('dept', '')
    sel_status = request.args.get('status', '')
    page       = request.args.get('page', 1, type=int)

    if not start_date and sel_date:
        start_date = sel_date
    if not start_date:
        start_date = today.isoformat()
    if not end_date:
        end_date = start_date

    logger.debug('Attendance filter — start_date=%s end_date=%s dept=%s status=%s',
                 start_date, end_date, sel_dept, sel_status)

    query = db.session.query(AttendanceLog, Employee)\
               .join(Employee, AttendanceLog.employee_id == Employee.id)\
               .filter(AttendanceLog.log_date >= start_date)\
               .filter(AttendanceLog.log_date <= end_date)
    if sel_dept:
        query = query.filter(Employee.department == sel_dept)
    if sel_status:
        query = query.filter(AttendanceLog.status == sel_status)
    pagination = query.order_by(Employee.department, Employee.full_name).paginate(page=page, per_page=20, error_out=False)
    logs = pagination.items

    depts = [d[0] for d in db.session.query(Employee.department).distinct().all()]
    return render_template('admin/attendance.html',
        logs=logs, sel_date=start_date, sel_dept=sel_dept,
        sel_status=sel_status, departments=depts, sel_end_date=end_date,
        work_hours_str=work_hours_str, page_obj=pagination)


@admin_attendance_bp.route('/admin/attendance/manual-clockin', methods=['POST'])
@admin_required
def admin_manual_clockin():
    data = request.get_json() or {}
    emp_id = data.get('employee_id')
    justification = (data.get('justification') or '').strip()
    if not emp_id or not justification:
        return jsonify({'ok': False, 'msg': 'يجب اختيار الموظف وإدخال مبرر.'}), 400
    emp = Employee.query.get(emp_id)
    if not emp:
        return jsonify({'ok': False, 'msg': 'الموظف غير موجود.'}), 404
    today = date.today()
    log = AttendanceLog.query.filter_by(employee_id=emp.id, log_date=today).first()
    if log and log.clock_in:
        return jsonify({'ok': False, 'msg': f'{emp.full_name} مسجل حضوره اليوم بالفعل.'})
    now = datetime.now()
    if log:
        log.clock_in = now; log.is_inside_geofence = True
        log.override_reason = justification
        log.status = 'present'
    else:
        log = AttendanceLog(employee_id=emp.id, log_date=today,
                            clock_in=now, is_inside_geofence=True,
                            override_reason=justification, status='present')
        db.session.add(log)
    n = Notification(employee_id=emp.id, title='تسجيل حضور يدوي',
        message=f'تم تسجيل حضورك يدوياً بواسطة الإدارة: {justification}',
        ntype='info', url='/employee')
    db.session.add(n)
    db.session.commit()
    return jsonify({'ok': True, 'msg': f'✓ تم تسجيل حضور {emp.full_name} يدوياً.'})





@admin_attendance_bp.route('/admin/reports/excel')
@admin_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    today  = date.today()
    month  = request.args.get('month', today.month, type=int)
    year   = request.args.get('year',  today.year,  type=int)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAMES[month-1]} {year}"
    ws.sheet_view.rightToLeft = True
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 35

    ws.merge_cells('A1:J1')
    title = ws['A1']
    title.value = f"تقرير الحضور والانصراف — {MONTH_NAMES[month-1]} {year} — SMARTLOG"
    title.font  = Font(bold=True, size=14, color='FFFFFF')
    title.fill  = PatternFill("solid", fgColor='991B1B')
    title.alignment = Alignment(horizontal='center', vertical='center')

    hdrs = ['#','الرقم الوظيفي','الاسم','القسم','أيام الحضور',
            'أيام التأخير','أيام الغياب','إجمالي التأخير(د)',
            'الراتب الأساسي','الخصم','صافي الراتب']
    hfill = PatternFill("solid", fgColor='DC2626')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor='FEF2F2')

    employees = Employee.query.filter_by(role='employee', is_active=True)\
                    .order_by(Employee.department, Employee.full_name)\
                    .options(db.joinedload(Employee.department_ref)).all()
    logs_all = AttendanceLog.query.filter(
        AttendanceLog.employee_id.in_([e.id for e in employees]),
        extract('month', AttendanceLog.log_date) == month,
        extract('year',  AttendanceLog.log_date) == year
    ).all()
    logs_by_emp = defaultdict(list)
    for l in logs_all:
        logs_by_emp[l.employee_id].append(l)
    for ri, emp in enumerate(employees, 3):
        logs = logs_by_emp.get(emp.id, [])
        p = sum(1 for l in logs if l.status in ('present','late'))
        lt= sum(1 for l in logs if l.status == 'late')
        ab= sum(1 for l in logs if l.status == 'absent')
        lm= sum(l.late_minutes for l in logs)
        ded= PayrollService.calculate_deduction(emp.base_salary, lm)
        row_data = [ri-2, emp.username, emp.full_name, emp.department,
                    p, lt, ab, lm, emp.base_salary or 0, ded,
                    round((emp.base_salary or 0)-ded, 2)]
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
            if fill:
                c.fill = fill

    from openpyxl.utils import get_column_letter
    for ci in range(1, len(hdrs)+1):
        try:
            mx = 10
            for r in range(1, min(len(employees)+4, 100)):
                v = str(ws.cell(row=r, column=ci).value or '')
                mx = max(mx, len(v))
            ws.column_dimensions[get_column_letter(ci)].width = max(12, min(30, mx+4))
        except Exception:
            pass

    out = BytesIO()
    wb.save(out); out.seek(0)
    fname = f'تقرير_الحضور_{MONTH_NAMES[month-1]}_{year}.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@admin_attendance_bp.route('/admin/reports/pdf')
@admin_required
def export_pdf():
    today  = date.today()
    month  = request.args.get('month', today.month, type=int)
    year   = request.args.get('year',  today.year,  type=int)
    dept   = request.args.get('dept',  '')

    qry = Employee.query.filter_by(role='employee', is_active=True)
    if dept: qry = qry.filter_by(department=dept)
    employees = qry.order_by(Employee.department, Employee.full_name).options(
        db.joinedload(Employee.department_ref)).all()
    emp_ids = [e.id for e in employees]
    logs_all = AttendanceLog.query.filter(
        AttendanceLog.employee_id.in_(emp_ids),
        extract('month', AttendanceLog.log_date) == month,
        extract('year',  AttendanceLog.log_date) == year
    ).all() if emp_ids else []
    logs_by_emp = defaultdict(list)
    for l in logs_all:
        logs_by_emp[l.employee_id].append(l)

    rows = []
    for emp in employees:
        logs = logs_by_emp.get(emp.id, [])
        p   = sum(1 for l in logs if l.status in ('present','late'))
        lt  = sum(1 for l in logs if l.status == 'late')
        ab  = sum(1 for l in logs if l.status == 'absent')
        lm  = sum(l.late_minutes for l in logs)
        ded = PayrollService.calculate_deduction(emp.base_salary, lm)
        rows.append({'username': emp.username, 'name': emp.full_name,
                     'dept': emp.department, 'present': p, 'late': lt,
                     'absent': ab, 'late_m': lm,
                     'salary': emp.base_salary or 0, 'ded': ded,
                     'net': round((emp.base_salary or 0)-ded, 2)})

    html = render_template('pdf/report.html',
        rows=rows, month_name=MONTH_NAMES[month-1],
        year=year, generated=datetime.now().strftime('%Y-%m-%d %H:%M'))

    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        from xhtml2pdf import pisa
        out = BytesIO()
        pisa.CreatePDF(html.encode('utf-8'), dest=out)
        out.seek(0)
        fname = f'تقرير_{MONTH_NAMES[month-1]}_{year}.pdf'
        return send_file(out, mimetype='application/pdf',
                         as_attachment=True, download_name=fname)
    except Exception:
        return make_response(html, 200, {'Content-Type': 'text/html; charset=utf-8'})


@admin_attendance_bp.route('/admin/reports/section')
@admin_required
def admin_section_report():
    today = date.today()
    departments = Department.query.filter_by(is_active=True).options(
        db.joinedload(Department.employees)).all()
    dept_emp_ids = defaultdict(list)
    emp_dept_map = {}
    emp_ids = []
    for e in Employee.query.with_entities(Employee.id, Employee.department).filter_by(is_active=True).yield_per(500):
        dept_emp_ids[e.department].append(e.id)
        emp_dept_map[e.id] = e.department
        emp_ids.append(e.id)
    today_logs = AttendanceLog.query.filter(
        AttendanceLog.log_date == today,
        AttendanceLog.employee_id.in_(emp_ids)
    ).with_entities(AttendanceLog.employee_id, AttendanceLog.status).all()
    dept_present = defaultdict(int)
    dept_late = defaultdict(int)
    for log_emp_id, log_status in ((l.employee_id, l.status) for l in today_logs):
        dept = emp_dept_map.get(log_emp_id)
        if not dept:
            continue
        if log_status in ('present', 'حاضر'):
            dept_present[dept] += 1
        elif log_status in ('late', 'متأخر'):
            dept_late[dept] += 1
    sections = []
    for dept in departments:
        emp_ids = dept_emp_ids.get(dept.name_ar, [])
        total = len(emp_ids)
        if total == 0:
            sections.append({'dept': dept, 'total_employees': 0, 'present_count': 0, 'late_count': 0, 'absent_count': 0, 'deficit_rate': 0.0})
            continue
        present = dept_present.get(dept.name_ar, 0)
        late = dept_late.get(dept.name_ar, 0)
        absent = total - (present + late)
        deficit_rate = round((absent / total) * 100, 2)
        sections.append({'dept': dept, 'total_employees': total, 'present_count': present, 'late_count': late, 'absent_count': absent, 'deficit_rate': deficit_rate})
    return render_template('admin/section_report.html', sections=sections, today=today)


@admin_attendance_bp.route('/admin/reports/employee/<int:eid>')
@admin_required
def employee_ledger(eid):
    emp   = Employee.query.get_or_404(eid)
    today = date.today()
    month = request.args.get('month', today.month, type=int)
    year  = request.args.get('year',  today.year,  type=int)
    logs = AttendanceLog.query.filter(
        AttendanceLog.employee_id == emp.id,
        extract('month', AttendanceLog.log_date) == month,
        extract('year',  AttendanceLog.log_date) == year
    ).order_by(AttendanceLog.log_date).all()
    shifts = ShiftSchedule.query.filter(
        ShiftSchedule.employee_id == emp.id,
        extract('month', ShiftSchedule.scheduled_date) == month,
        extract('year',  ShiftSchedule.scheduled_date) == year
    ).all()
    total_shifts = len(shifts) or 30
    shifts_attended = sum(1 for l in logs if l.status in ('present', 'late'))
    total_late_minutes = sum(l.late_minutes for l in logs)
    absent_count = sum(1 for l in logs if l.status == 'absent')
    base = round(emp.base_salary or 0, 2)
    deduction_late = PayrollService.calculate_deduction(base, total_late_minutes)
    deduction_absent = PayrollService.calculate_deduction(base, absent_count * 8 * 60)
    total_deductions = round(deduction_late + deduction_absent, 2)
    net_salary = round(base - total_deductions, 2)
    attendance_rows = [{
        'date': l.log_date, 'clock_in': l.clock_in, 'clock_out': l.clock_out,
        'status': l.status, 'late_minutes': l.late_minutes,
        'is_inside_geofence': l.is_inside_geofence
    } for l in logs]
    summary = {'base': base, 'deductions': total_deductions, 'net': net_salary, 'minutes': total_late_minutes}
    return render_template('admin/employee_ledger.html',
        employee=emp, attendance_rows=attendance_rows, summary=summary,
        total_shifts=total_shifts, shifts_attended=shifts_attended,
        month=month, year=year, month_name=MONTH_NAMES[month-1], months=MONTH_NAMES)
