import json, csv, io, re, logging
from datetime import datetime, date, UTC
from functools import wraps

from flask import Blueprint, request, jsonify, session, send_file
from sqlalchemy import func

LOGGER = logging.getLogger(__name__)


def safe_json(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except Exception as e:
            LOGGER.error('API error in %s: %s', f.__name__, e)
            return jsonify({'ok': False, 'msg': str(e)}), 500
    return wrapper

from models import db
from models.employee import Employee
from models.department import Department, DepartmentCertification, DepartmentAnnouncement, DepartmentTransfer
from models.anomaly import AttendanceAnomaly
from models.documents import ArchivedDocument
from models.notifications import Notification
from models.biotime_device import BioTimeDevice
from models.shifts import ShiftType
from models.attendance import AttendanceLog

admin_departments_bp = Blueprint('admin_departments', __name__, url_prefix='/admin/departments')

DEPARTMENT_ICONS = ['flask', 'droplet', 'users', 'wallet', 'tool', 'clipboard-list', 'truck-medical', 'microscope', 'box', 'shield', 'heart-pulse', 'hospital', 'stethoscope', 'syringe', 'bandage', 'file-medical', 'chart-bar', 'clock', 'calendar-check', 'building']

def serialize_department(d):
    data = d.to_dict()
    data['certifications'] = [c.certification for c in d.required_certifications]
    data['allowed_device_ids'] = [dev.id for dev in d.allowed_devices]
    data['alert_recipient_ids'] = [emp.id for emp in d.alert_recipients]
    return data

def can_manage_departments():
    role = session.get('role')
    return role in ('admin', 'super_admin', 'hr_manager')

@admin_departments_bp.before_request
def require_department_manager():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403

@admin_departments_bp.route('/api/list')
@safe_json
def list_departments():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    depts = Department.query.order_by(Department.id).all()
    return jsonify({'departments': [serialize_department(d) for d in depts]})

@admin_departments_bp.route('/api/tree')
@safe_json
def department_tree():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    depts = Department.query.order_by(Department.id).all()
    tree = []
    for d in depts:
        if not d.parent_id:
            tree.append(build_tree_node(d))
    return jsonify({'tree': tree})

def build_tree_node(d):
    node = {
        'id': d.id,
        'code': d.code,
        'name_ar': d.name_ar,
        'name_en': d.name_en,
        'icon': d.icon,
        'color': d.color,
        'dept_type': d.dept_type,
        'manager_name': d.manager.full_name if d.manager else None,
        'employee_count': d.employee_count,
        'children': [build_tree_node(c) for c in d.children if c.is_active],
    }
    return node

@admin_departments_bp.route('/api/<int:dept_id>')
@safe_json
def get_department(dept_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    d = Department.query.get_or_404(dept_id)
    return jsonify(serialize_department(d))

@admin_departments_bp.route('/api/check-code')
@safe_json
def check_code():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    code = request.args.get('code', '').strip()
    if not code:
        return jsonify({'available': False, 'message': 'الرجاء إدخال كود'})
    existing = Department.query.filter_by(code=code).first()
    return jsonify({'available': not bool(existing), 'message': 'الكود متاح' if not existing else 'الكود مستخدم بالفعل'})

@admin_departments_bp.route('/api/generate-code')
@safe_json
def generate_code():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    return jsonify({'code': Department.generate_code()})

@admin_departments_bp.route('/api/parents')
@safe_json
def list_parents():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    depts = Department.query.order_by(Department.id).all()
    return jsonify({'parents': [{'id': d.id, 'name_ar': d.name_ar, 'hierarchy_path': d.hierarchy_path} for d in depts]})

@admin_departments_bp.route('/api/managers')
@safe_json
def list_managers():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    query = request.args.get('q', '').strip()
    q = Employee.query.filter_by(deleted_at=None, is_active=True)
    if query:
        q = q.filter(Employee.full_name.contains(query))
    emps = q.order_by(Employee.full_name).limit(20).all()
    return jsonify({'employees': [{'id': e.id, 'full_name': e.full_name, 'username': e.username, 'profile_photo': e.profile_photo} for e in emps]})

@admin_departments_bp.route('/api/devices')
@safe_json
def list_devices():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    devices = BioTimeDevice.query.filter_by(is_active=True).all()
    return jsonify({'devices': [{'id': d.id, 'name': d.name or d.serial_number, 'serial_number': d.serial_number} for d in devices]})

@admin_departments_bp.route('/api/shifts')
@safe_json
def list_shifts():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    shifts = ShiftType.query.filter_by(is_active=True).all()
    return jsonify({'shifts': [{'id': s.id, 'name_ar': s.name_ar} for s in shifts]})

@admin_departments_bp.route('/api/employees')
@safe_json
def list_employees():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    query = request.args.get('q', '').strip()
    q = Employee.query.filter_by(deleted_at=None, is_active=True)
    if query:
        q = q.filter(Employee.full_name.contains(query))
    emps = q.order_by(Employee.full_name).limit(20).all()
    return jsonify({'employees': [{'id': e.id, 'full_name': e.full_name, 'username': e.username} for e in emps]})

@admin_departments_bp.route('/api/add', methods=['POST'])
@safe_json
def add_department():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json() or {}
    code = data.get('code', '').strip()
    name_ar = data.get('name_ar', '').strip()
    if not code:
        code = Department.generate_code()
    if not name_ar:
        return jsonify({'error': 'اسم القسم مطلوب'}), 400
    existing = Department.query.filter_by(code=code).first()
    if existing:
        return jsonify({'error': f'الكود {code} مستخدم بالفعل'}), 400
    parent_id = data.get('parent_id')
    if parent_id:
        parent = Department.query.get(int(parent_id))
        dept_level = (parent.dept_level or 0) + 1
    else:
        parent_id = None
        dept_level = 1
    d = Department(
        code=code,
        name_ar=name_ar,
        name_en=data.get('name_en', '').strip() or None,
        icon=data.get('icon', 'building'),
        color=data.get('color', '#e53935'),
        description_ar=data.get('description_ar', '').strip() or None,
        description_en=data.get('description_en', '').strip() or None,
        dept_type=data.get('dept_type', 'operational'),
        is_active=data.get('is_active', True),
        inactive_reason=data.get('inactive_reason', '').strip() or None,
        parent_id=parent_id,
        dept_level=dept_level,
        manager_id=int(data['manager_id']) if data.get('manager_id') else None,
        deputy_id=int(data['deputy_id']) if data.get('deputy_id') else None,
        cost_center_code=data.get('cost_center_code', '').strip() or None,
        min_staff_required=int(data.get('min_staff_required', 2)),
        max_staff_capacity=int(data.get('max_staff_capacity', 50)),
        allowed_employment_types=data.get('allowed_employment_types', 'full_time,part_time'),
        default_shift_id=int(data['default_shift_id']) if data.get('default_shift_id') else None,
        grace_period_override=int(data['grace_period_override']) if data.get('grace_period_override') else None,
        remote_work_allowed=bool(data.get('remote_work_allowed', False)),
        break_duration_policy=int(data.get('break_duration_policy', 60)),
        overtime_max_weekly=int(data.get('overtime_max_weekly', 12)),
        overtime_requires_approval=bool(data.get('overtime_requires_approval', True)),
        overtime_auto_approve_under=int(data.get('overtime_auto_approve_under', 2)),
        whatsapp_group_id=data.get('whatsapp_group_id', '').strip() or None,
        alert_threshold_minutes=int(data.get('alert_threshold_minutes', 15)),
        alert_understaffing_threshold=int(data['alert_understaffing_threshold']) if data.get('alert_understaffing_threshold') else None,
    )
    if data.get('alert_settings'):
        d.alert_settings_dict = data['alert_settings']
    db.session.add(d)
    db.session.flush()
    certs = data.get('certifications', [])
    for cert in certs:
        cert_name = cert.strip() if isinstance(cert, str) else cert.get('name', '').strip()
        if cert_name:
            dc = DepartmentCertification(department_id=d.id, certification=cert_name)
            db.session.add(dc)
    device_ids = data.get('allowed_device_ids', [])
    if device_ids:
        devices = BioTimeDevice.query.filter(BioTimeDevice.id.in_(device_ids)).all()
        d.allowed_devices = devices
    recipient_ids = data.get('alert_recipient_ids', [])
    if recipient_ids:
        recipients = Employee.query.filter(Employee.id.in_(recipient_ids)).all()
        d.alert_recipients = recipients
    db.session.commit()
    return jsonify({'success': True, 'department': serialize_department(d)})

@admin_departments_bp.route('/api/<int:dept_id>/edit', methods=['POST'])
@safe_json
def edit_department(dept_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    d = Department.query.get_or_404(dept_id)
    data = request.get_json() or {}
    name_ar = data.get('name_ar', '').strip()
    if not name_ar:
        return jsonify({'error': 'اسم القسم مطلوب'}), 400
    new_code = data.get('code', '').strip()
    if new_code and new_code != d.code:
        existing = Department.query.filter_by(code=new_code).first()
        if existing:
            return jsonify({'error': f'الكود {new_code} مستخدم بالفعل'}), 400
        d.code = new_code
    d.name_ar = name_ar
    d.name_en = data.get('name_en', '').strip() or None
    d.icon = data.get('icon', d.icon)
    d.color = data.get('color', d.color)
    d.description_ar = data.get('description_ar', '').strip() or None
    d.description_en = data.get('description_en', '').strip() or None
    d.dept_type = data.get('dept_type', d.dept_type)
    is_active = data.get('is_active')
    if is_active is not None:
        d.is_active = bool(is_active)
    d.inactive_reason = data.get('inactive_reason', '').strip() or None
    new_parent_id = data.get('parent_id')
    if new_parent_id is not None:
        if int(new_parent_id) == d.id:
            return jsonify({'error': 'لا يمكن جعل القسم تابعاً لنفسه'}), 400
        d.parent_id = int(new_parent_id) if new_parent_id else None
        if d.parent_id:
            parent = Department.query.get(d.parent_id)
            d.dept_level = (parent.dept_level or 0) + 1
        else:
            d.dept_level = 1
    d.manager_id = int(data['manager_id']) if data.get('manager_id') else None
    d.deputy_id = int(data['deputy_id']) if data.get('deputy_id') else None
    d.cost_center_code = data.get('cost_center_code', '').strip() or None
    d.min_staff_required = int(data.get('min_staff_required', d.min_staff_required))
    d.max_staff_capacity = int(data.get('max_staff_capacity', d.max_staff_capacity))
    d.allowed_employment_types = data.get('allowed_employment_types', d.allowed_employment_types)
    d.default_shift_id = int(data['default_shift_id']) if data.get('default_shift_id') else None
    d.grace_period_override = int(data['grace_period_override']) if data.get('grace_period_override') else None
    d.remote_work_allowed = bool(data.get('remote_work_allowed', d.remote_work_allowed))
    d.break_duration_policy = int(data.get('break_duration_policy', d.break_duration_policy))
    d.overtime_max_weekly = int(data.get('overtime_max_weekly', d.overtime_max_weekly))
    d.overtime_requires_approval = bool(data.get('overtime_requires_approval', d.overtime_requires_approval))
    d.overtime_auto_approve_under = int(data.get('overtime_auto_approve_under', d.overtime_auto_approve_under))
    if 'whatsapp_group_id' in data:
        d.whatsapp_group_id = data.get('whatsapp_group_id', '').strip() or None
    if data.get('alert_settings'):
        d.alert_settings_dict = data['alert_settings']
    if 'alert_threshold_minutes' in data:
        d.alert_threshold_minutes = int(data['alert_threshold_minutes'])
    if 'alert_understaffing_threshold' in data:
        d.alert_understaffing_threshold = int(data['alert_understaffing_threshold']) if data.get('alert_understaffing_threshold') else None
    if 'certifications' in data:
        DepartmentCertification.query.filter_by(department_id=d.id).delete()
        for cert in data['certifications']:
            cert_name = cert.strip() if isinstance(cert, str) else cert.get('name', '').strip()
            if cert_name:
                dc = DepartmentCertification(department_id=d.id, certification=cert_name)
                db.session.add(dc)
    if 'allowed_device_ids' in data:
        device_ids = data['allowed_device_ids']
        d.allowed_devices = BioTimeDevice.query.filter(BioTimeDevice.id.in_(device_ids)).all() if device_ids else []
    if 'alert_recipient_ids' in data:
        recipient_ids = data['alert_recipient_ids']
        d.alert_recipients = Employee.query.filter(Employee.id.in_(recipient_ids)).all() if recipient_ids else []
    d.updated_at = datetime.now(UTC)
    db.session.commit()
    return jsonify({'success': True, 'department': serialize_department(d)})

@admin_departments_bp.route('/api/<int:dept_id>/delete', methods=['POST'])
@safe_json
def delete_department(dept_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    d = Department.query.get_or_404(dept_id)
    emp_count = d.employee_count
    if emp_count > 0:
        return jsonify({'error': f'لا يمكن حذف القسم — لديه {emp_count} موظف', 'employee_count': emp_count}), 400
    child_count = Department.query.filter_by(parent_id=dept_id).count()
    if child_count > 0:
        return jsonify({'error': f'لا يمكن حذف القسم — لديه {child_count} قسم فرعي', 'child_count': child_count}), 400
    DepartmentCertification.query.filter_by(department_id=d.id).delete()
    DepartmentAnnouncement.query.filter_by(department_id=d.id).delete()
    DepartmentTransfer.query.filter_by(from_department_id=dept_id).update({'from_department_id': None})
    DepartmentTransfer.query.filter_by(to_department_id=dept_id).update({'to_department_id': None})
    db.session.delete(d)
    db.session.commit()
    return jsonify({'success': True})

@admin_departments_bp.route('/api/<int:dept_id>/toggle', methods=['POST'])
@safe_json
def toggle_department(dept_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    d = Department.query.get_or_404(dept_id)
    data = request.get_json() or {}
    d.is_active = not d.is_active
    if not d.is_active:
        d.inactive_reason = data.get('reason', '').strip() or 'تم الإيقاف'
    else:
        d.inactive_reason = None
    d.updated_at = datetime.now(UTC)
    db.session.commit()
    return jsonify({'success': True, 'is_active': d.is_active})

@admin_departments_bp.route('/api/bulk-toggle', methods=['POST'])
@safe_json
def bulk_toggle():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json() or {}
    dept_ids = data.get('department_ids', [])
    is_active = bool(data.get('is_active', True))
    reason = data.get('reason', '').strip()
    count = 0
    for did in dept_ids:
        d = Department.query.get(int(did))
        if d:
            d.is_active = is_active
            d.inactive_reason = reason if not is_active else None
            d.updated_at = datetime.now(UTC)
            count += 1
    db.session.commit()
    return jsonify({'success': True, 'updated': count})

@admin_departments_bp.route('/api/<int:dept_id>/clone', methods=['POST'])
@safe_json
def clone_department(dept_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    source = Department.query.get_or_404(dept_id)
    data = request.get_json() or {}
    new_name = data.get('name_ar', '').strip() or f'{source.name_ar} (نسخة)'
    new_code = Department.generate_code()
    d = Department(
        code=new_code,
        name_ar=new_name,
        name_en=f'{source.name_en} (Copy)' if source.name_en else None,
        icon=source.icon,
        color=source.color,
        description_ar=source.description_ar,
        description_en=source.description_en,
        dept_type=source.dept_type,
        is_active=True,
        parent_id=source.parent_id,
        dept_level=source.dept_level,
        manager_id=None,
        deputy_id=None,
        cost_center_code=None,
        min_staff_required=source.min_staff_required,
        max_staff_capacity=source.max_staff_capacity,
        allowed_employment_types=source.allowed_employment_types,
        default_shift_id=source.default_shift_id,
        grace_period_override=source.grace_period_override,
        remote_work_allowed=source.remote_work_allowed,
        break_duration_policy=source.break_duration_policy,
        overtime_max_weekly=source.overtime_max_weekly,
        overtime_requires_approval=source.overtime_requires_approval,
        overtime_auto_approve_under=source.overtime_auto_approve_under,
        whatsapp_group_id=None,
        alert_threshold_minutes=source.alert_threshold_minutes,
        alert_understaffing_threshold=source.alert_understaffing_threshold,
    )
    if source.alert_settings:
        d.alert_settings = source.alert_settings
    db.session.add(d)
    db.session.flush()
    for cert in source.required_certifications:
        dc = DepartmentCertification(department_id=d.id, certification=cert.certification)
        db.session.add(dc)
    d.allowed_devices = list(source.allowed_devices)
    d.alert_recipients = list(source.alert_recipients)
    db.session.commit()
    return jsonify({'success': True, 'department': serialize_department(d)})

@admin_departments_bp.route('/api/<int:dept_id>/dashboard')
@safe_json
def department_dashboard(dept_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    d = Department.query.get_or_404(dept_id)
    today = date.today()
    employees = Employee.query.filter_by(department_id=dept_id, deleted_at=None, is_active=True).all()
    emp_ids = [e.id for e in employees]
    total_emp = len(emp_ids)
    today_logs = AttendanceLog.query.filter(
        AttendanceLog.employee_id.in_(emp_ids) if emp_ids else False,
        AttendanceLog.log_date == today,
    ).all() if emp_ids else []
    clocked_ids = set(l.employee_id for l in today_logs if l.clock_in)
    present = len(clocked_ids)
    absent = total_emp - present
    late = 0
    for log in today_logs:
        if log.clock_in and log.clock_in.hour >= 9:
            late += 1
    month_start = today.replace(day=1)
    month_logs = AttendanceLog.query.filter(
        AttendanceLog.employee_id.in_(emp_ids) if emp_ids else False,
        AttendanceLog.log_date >= month_start,
        AttendanceLog.log_date <= today,
        AttendanceLog.clock_in.isnot(None),
    ).all() if emp_ids else []
    month_present_days = len(set((l.employee_id, l.log_date) for l in month_logs))
    month_total_days = total_emp * ((today - month_start).days + 1)
    month_attendance_pct = round((month_present_days / month_total_days) * 100, 1) if month_total_days > 0 else 0
    top_punctual = db.session.query(
        Employee.id, Employee.full_name, Employee.profile_photo,
        func.count(AttendanceLog.id).label('punch_count')
    ).join(AttendanceLog).filter(
        AttendanceLog.employee_id.in_(emp_ids) if emp_ids else False,
        AttendanceLog.log_date >= month_start,
        AttendanceLog.clock_in.isnot(None),
    ).group_by(Employee.id).order_by(func.count(AttendanceLog.id).desc()).limit(3).all() if emp_ids else []
    pending_leaves = 0
    from models.misc import LeaveRequest
    pending_leaves = LeaveRequest.query.filter(
        LeaveRequest.employee_id.in_(emp_ids) if emp_ids else False,
        LeaveRequest.status == 'pending',
    ).count() if emp_ids else 0
    expiring_docs = ArchivedDocument.query.filter(
        ArchivedDocument.employee_id.in_(emp_ids) if emp_ids else False,
        ArchivedDocument.expiry_date.isnot(None),
        ArchivedDocument.expiry_date <= today,
        ArchivedDocument.expiry_date >= today,
        ArchivedDocument.is_deleted == False,
    ).count() if emp_ids else 0
    active_anomalies = 0
    try:
        active_anomalies = AttendanceAnomaly.query.filter(
            AttendanceAnomaly.employee_id.in_(emp_ids) if emp_ids else False,
            AttendanceAnomaly.status == 'open',
        ).count() if emp_ids else 0
    except Exception:
        active_anomalies = 0
    attendance_rate = round((present / total_emp) * 100, 1) if total_emp > 0 else 0
    return jsonify({
        'total_employees': total_emp,
        'present': present,
        'absent': absent,
        'late': late,
        'attendance_rate': attendance_rate,
        'month_attendance_pct': month_attendance_pct,
        'top_punctual': [{'id': e.id, 'full_name': e.full_name, 'profile_photo': e.profile_photo} for e in top_punctual],
        'pending_leaves': pending_leaves,
        'expiring_documents': expiring_docs,
        'active_anomalies': active_anomalies,
    })

@admin_departments_bp.route('/api/<int:dept_id>/announcements', methods=['GET'])
@safe_json
def list_announcements(dept_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    dept = Department.query.get_or_404(dept_id)
    announcements = DepartmentAnnouncement.query.filter_by(department_id=dept_id).order_by(DepartmentAnnouncement.created_at.desc()).all()
    return jsonify({'announcements': [a.to_dict() for a in announcements]})

@admin_departments_bp.route('/api/<int:dept_id>/announcements/send', methods=['POST'])
@safe_json
def send_announcement(dept_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json() or {}
    message = data.get('message', '').strip()
    if not message:
        return jsonify({'error': 'الرسالة مطلوبة'}), 400
    delivery_methods = data.get('delivery_method', 'in_app')
    if isinstance(delivery_methods, list):
        delivery_methods = ','.join(delivery_methods)
    priority = data.get('priority', 'normal')
    scheduled_at = None
    if data.get('schedule_later') and data.get('scheduled_at'):
        try:
            scheduled_at = datetime.fromisoformat(data['scheduled_at'])
        except (ValueError, TypeError):
            pass
    target_type = data.get('target_type', 'all')
    target_ids = None
    if target_type == 'specific' and data.get('target_employee_ids'):
        target_ids = json.dumps(data['target_employee_ids'], ensure_ascii=False)
    ann = DepartmentAnnouncement(
        department_id=dept_id,
        message=message,
        priority=priority,
        delivery_method=delivery_methods,
        scheduled_at=scheduled_at,
        sent_by=session.get('user_id'),
        target_type=target_type,
        target_ids=target_ids,
    )
    if not scheduled_at or scheduled_at <= datetime.now(UTC):
        ann.sent_at = datetime.now(UTC)
    db.session.add(ann)
    db.session.flush()
    employees = Employee.query.filter_by(department_id=dept_id, deleted_at=None, is_active=True).all()
    if target_type == 'specific' and data.get('target_employee_ids'):
        employees = Employee.query.filter(Employee.id.in_(data['target_employee_ids'])).all()
    if 'in_app' in delivery_methods or 'in_app' in (delivery_methods or ''):
        for emp in employees:
            notif = Notification(
                employee_id=emp.id,
                title='إعلان قسم' if priority == 'normal' else ('إعلان عاجل' if priority == 'urgent' else 'إعلان مهم جداً'),
                message=re.sub(r'<[^>]*>', '', ann.message),
                ntype='announcement',
                is_read=False,
            )
            db.session.add(notif)
    db.session.commit()
    total_target = len(employees)
    return jsonify({
        'success': True,
        'announcement': ann.to_dict(),
        'sent_to': total_target,
        'message': f'تم إرسال الإعلان لـ {total_target} موظف',
    })

@admin_departments_bp.route('/api/transfers', methods=['GET'])
@safe_json
def list_transfers():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    transfers = DepartmentTransfer.query.order_by(DepartmentTransfer.created_at.desc()).limit(50).all()
    return jsonify({'transfers': [t.to_dict() for t in transfers]})

@admin_departments_bp.route('/api/transfers/create', methods=['POST'])
@safe_json
def create_transfer():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json() or {}
    employee_id = int(data.get('employee_id', 0))
    to_department_id = int(data.get('to_department_id', 0))
    if not employee_id or not to_department_id:
        return jsonify({'error': 'الموظف والقسم المطلوب إلزاميان'}), 400
    employee = Employee.query.get(employee_id)
    if not employee:
        return jsonify({'error': 'الموظف غير موجود'}), 404
    to_dept = Department.query.get(to_department_id)
    if not to_dept:
        return jsonify({'error': 'القسم غير موجود'}), 404
    transfer_date = date.today()
    if data.get('transfer_date'):
        try:
            transfer_date = date.fromisoformat(data['transfer_date'])
        except (ValueError, TypeError):
            pass
    t = DepartmentTransfer(
        employee_id=employee_id,
        from_department_id=employee.department_id,
        to_department_id=to_department_id,
        transfer_date=transfer_date,
        reason_type=data.get('reason_type', ''),
        reason_notes=data.get('reason_notes', ''),
    )
    db.session.add(t)
    db.session.commit()
    return jsonify({'success': True, 'transfer': t.to_dict()})

@admin_departments_bp.route('/api/transfers/<int:t_id>/approve-manager', methods=['POST'])
@safe_json
def approve_transfer_manager(t_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    t = DepartmentTransfer.query.get_or_404(t_id)
    t.approved_by_manager = True
    t.manager_approved_at = datetime.now(UTC)
    if t.approved_by_hr:
        t.status = 'approved'
    db.session.commit()
    return jsonify({'success': True, 'transfer': t.to_dict()})

@admin_departments_bp.route('/api/transfers/<int:t_id>/approve-hr', methods=['POST'])
@safe_json
def approve_transfer_hr(t_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    t = DepartmentTransfer.query.get_or_404(t_id)
    t.approved_by_hr = True
    t.hr_approved_at = datetime.now(UTC)
    if t.approved_by_manager:
        t.status = 'approved'
    db.session.commit()
    return jsonify({'success': True, 'transfer': t.to_dict()})

@admin_departments_bp.route('/api/transfers/<int:t_id>/execute', methods=['POST'])
@safe_json
def execute_transfer(t_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    t = DepartmentTransfer.query.get_or_404(t_id)
    if t.status != 'approved':
        return jsonify({'error': 'يجب الموافقة على النقل أولاً'}), 400
    employee = Employee.query.get(t.employee_id)
    if not employee:
        return jsonify({'error': 'الموظف غير موجود'}), 404
    employee.department_id = t.to_department_id
    employee.department = t.to_department.name_ar
    t.status = 'executed'
    t.executed_at = datetime.now(UTC)
    db.session.commit()
    return jsonify({'success': True, 'transfer': t.to_dict()})

@admin_departments_bp.route('/api/transfers/<int:t_id>/reject', methods=['POST'])
@safe_json
def reject_transfer(t_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    t = DepartmentTransfer.query.get_or_404(t_id)
    t.status = 'rejected'
    db.session.commit()
    return jsonify({'success': True, 'transfer': t.to_dict()})

@admin_departments_bp.route('/api/export')
def export_departments():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    depts = Department.query.order_by(Department.id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'departments'
    ws.sheet_view.rightToLeft = True
    hdrs = ['code', 'name_ar', 'name_en', 'dept_type', 'parent_code', 'manager_id', 'min_staff_required', 'max_staff_capacity', 'default_shift_id', 'is_active']
    hfill = PatternFill('solid', fgColor='DC2626')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = Alignment(horizontal='center')
    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill('solid', fgColor='FEF2F2')
    for ri, d in enumerate(depts, 2):
        parent_code = d.parent.code if d.parent else ''
        vals = [d.code, d.name_ar, d.name_en or '', d.dept_type, parent_code, d.manager_id or '', d.min_staff_required, d.max_staff_capacity, d.default_shift_id or '', '1' if d.is_active else '0']
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border; c.alignment = Alignment(horizontal='center')
            if ri % 2 == 0: c.fill = alt_fill
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'departments_{date.today().isoformat()}.xlsx')

@admin_departments_bp.route('/api/import-template')
def import_template():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(['code', 'name_ar', 'name_en', 'dept_type', 'parent_code', 'manager_id', 'min_staff_required', 'max_staff_capacity', 'default_shift_id', 'is_active'])
    writer.writerow(['DEPT-010', 'قسم جديد', 'New Department', 'operational', 'DEPT-001', '1', '2', '50', '1', '1'])
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='import_template.csv',
    )

@admin_departments_bp.route('/api/import', methods=['POST'])
@safe_json
def import_departments():
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'الرجاء رفع ملف CSV'}), 400
    try:
        content = file.read().decode('utf-8-sig').strip()
        reader = csv.DictReader(io.StringIO(content))
    except Exception as e:
        return jsonify({'error': f'خطأ في قراءة الملف: {str(e)}'}), 400
    imported = 0
    errors = []
    for row_num, row in enumerate(reader, start=2):
        code = row.get('code', '').strip()
        name_ar = row.get('name_ar', '').strip()
        if not name_ar:
            errors.append(f'الصف {row_num}: اسم القسم مطلوب')
            continue
        if not code:
            code = Department.generate_code()
        if Department.query.filter_by(code=code).first():
            errors.append(f'الصف {row_num}: الكود {code} موجود بالفعل')
            continue
        parent_code = row.get('parent_code', '').strip()
        parent_id = None
        if parent_code:
            parent = Department.query.filter_by(code=parent_code).first()
            if parent:
                parent_id = parent.id
        manager_id = None
        try:
            manager_id = int(row.get('manager_id', '')) if row.get('manager_id', '').strip() else None
        except ValueError:
            pass
        shift_id = None
        try:
            shift_id = int(row.get('default_shift_id', '')) if row.get('default_shift_id', '').strip() else None
        except ValueError:
            pass
        min_staff = 2
        try:
            min_staff = int(row.get('min_staff_required', '2'))
        except ValueError:
            pass
        max_cap = 50
        try:
            max_cap = int(row.get('max_staff_capacity', '50'))
        except ValueError:
            pass
        is_active = row.get('is_active', '1') == '1'
        dept_type = row.get('dept_type', 'operational').strip()
        if dept_type not in ('operational', 'administrative', 'medical', 'technical', 'support'):
            dept_type = 'operational'
        d = Department(
            code=code,
            name_ar=name_ar,
            name_en=row.get('name_en', '').strip() or None,
            dept_type=dept_type,
            parent_id=parent_id,
            manager_id=manager_id,
            min_staff_required=min_staff,
            max_staff_capacity=max_cap,
            default_shift_id=shift_id,
            is_active=is_active,
        )
        db.session.add(d)
        imported += 1
    db.session.commit()
    return jsonify({'success': True, 'imported': imported, 'errors': errors})

@admin_departments_bp.route('/api/<int:dept_id>/employees')
@safe_json
def dept_employees(dept_id):
    if not can_manage_departments():
        return jsonify({'error': 'Unauthorized'}), 403
    department = Department.query.get_or_404(dept_id)
    employees = Employee.query.filter_by(department_id=dept_id, deleted_at=None, is_active=True).all()
    return jsonify({'employees': [{
        'id': e.id,
        'username': e.username,
        'full_name': e.full_name,
        'job_title': e.job_title,
        'profile_photo': e.profile_photo,
        'phone': e.phone,
        'employment_type': e.employment_type,
    } for e in employees]})

@admin_departments_bp.route('/api/icons')
@safe_json
def list_icons():
    return jsonify({'icons': DEPARTMENT_ICONS})
