import html, json, math, logging
from datetime import datetime, date, timedelta, UTC
from collections import defaultdict
from functools import wraps

from flask import Blueprint, request, jsonify, render_template, session
from sqlalchemy import func, extract

from models import db
from models.employee import Employee
from models.department import Department
from models.attendance import AttendanceLog
from models.misc import LeaveRequest, EmployeeDocument
from models.misc import EmployeeDocument
from models.biotime_device import BioTimeDevice
from models.shifts import ShiftType
from models.notifications import Notification
from models.employee_enhanced import EmployeeExtended, EmployeeLeaveRequest as NewLeaveRequest, EmployeePromotion, EmployeeGrade
from utils.decorators import admin_required

admin_dashboard_bp = Blueprint('admin_dashboard', __name__)
LOGGER = logging.getLogger(__name__)


def safe_json_response(f):
    """Wrap API endpoints with try/except that always returns JSON."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except Exception as e:
            LOGGER.error('API error in %s: %s', f.__name__, e)
            return jsonify({'ok': False, 'msg': str(e), 'data': []}), 500
    return wrapper

DAY_NAMES = ['الاثنين','الثلاثاء','الأربعاء','الخميس','الجمعة','السبت','الأحد']
MONTH_NAMES = ['','يناير','فبراير','مارس','أبريل','مايو','يونيو','يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر']


@admin_dashboard_bp.route('/admin')
@admin_required
def admin_dashboard():
    today = date.today()
    return render_template('admin/dashboard.html',
        today=today, month_name=f'{today.month:02d}-{today.year}',
        day_name=DAY_NAMES[today.weekday()])


@admin_dashboard_bp.route('/api/dashboard/stats')
@safe_json_response
@admin_required
def api_dashboard_stats():
    today = date.today()
    yesterday = today - timedelta(days=1)
    total = Employee.query.filter_by(deleted_at=None, is_active=True).count()
    t_logs = AttendanceLog.query.filter_by(log_date=today).all()
    y_logs = AttendanceLog.query.filter_by(log_date=yesterday).all()
    present = sum(1 for l in t_logs if l.status in ('present', 'late'))
    late = sum(1 for l in t_logs if l.status == 'late')
    absent = total - present
    on_leave = LeaveRequest.query.filter(
        LeaveRequest.status == 'approved',
        LeaveRequest.start_date <= today,
        LeaveRequest.end_date >= today,
    ).count()
    no_clockout = sum(1 for l in t_logs if l.clock_in and not l.clock_out and l.status != 'absent')
    expiring_docs = EmployeeDocument.query.filter(
        EmployeeDocument.expiry_date.isnot(None),
        EmployeeDocument.expiry_date <= today + timedelta(days=30),
        EmployeeDocument.expiry_date >= today,
    ).count()
    offline_devices = BioTimeDevice.query.filter_by(is_active=True).filter(
        BioTimeDevice.is_online == False
    ).count()
    pending_leave_requests = NewLeaveRequest.query.filter_by(status='pending').count()
    total_with_extended = EmployeeExtended.query.count()
    total_emps = Employee.query.filter_by(deleted_at=None, is_active=True).count()
    extended_pct = round(total_with_extended / total_emps * 100) if total_emps else 0
    y_present = sum(1 for l in y_logs if l.status in ('present', 'late'))
    y_absent = total - y_present
    y_late = sum(1 for l in y_logs if l.status == 'late')
    get_trend = lambda curr, prev: (curr - prev) if prev else 0
    return jsonify({
        'total': total,
        'present': present,
        'absent': absent,
        'late': late,
        'on_leave': on_leave,
        'no_clockout': no_clockout,
        'expiring_docs': expiring_docs,
        'offline_devices': offline_devices,
        'pending_leave_requests': pending_leave_requests,
        'extended_data_pct': extended_pct,
        'trends': {
            'present': get_trend(present, y_present),
            'absent': get_trend(absent, y_absent),
            'late': get_trend(late, y_late),
        },
        'total_employees_no_leave': total - on_leave,
    })


@admin_dashboard_bp.route('/api/dashboard/charts/weekly')
@safe_json_response
@admin_required
def api_charts_weekly():
    today = date.today()
    mode = request.args.get('mode', 'weekly')
    total = Employee.query.filter_by(deleted_at=None, is_active=True).count()
    on_leave_ids = set()
    leave_reqs = LeaveRequest.query.filter(
        LeaveRequest.status == 'approved',
    ).all()
    data = []
    if mode == 'weekly':
        days = 7
        date_range = [today - timedelta(days=i) for i in range(days - 1, -1, -1)]
    elif mode == 'monthly':
        days = (today - today.replace(day=1)).days + 1
        date_range = [today - timedelta(days=i) for i in range(days - 1, -1, -1)]
    else:
        days = 30
        date_range = [today - timedelta(days=i) for i in range(days - 1, -1, -1)]
    for d in date_range:
        logs = AttendanceLog.query.filter_by(log_date=d).all()
        day_present = sum(1 for l in logs if l.status in ('present', 'late'))
        day_late = sum(1 for l in logs if l.status == 'late')
        day_absent = total - day_present
        label = DAY_NAMES[d.weekday()] if mode == 'weekly' else d.strftime('%d/%m')
        data.append({
            'day': label,
            'present': day_present,
            'absent': day_absent,
            'late': day_late,
        })
    return jsonify({'data': data, 'mode': mode, 'total': total})


@admin_dashboard_bp.route('/api/dashboard/charts/donut')
@safe_json_response
@admin_required
def api_charts_donut():
    today = date.today()
    total = Employee.query.filter_by(deleted_at=None, is_active=True).count()
    t_logs = AttendanceLog.query.filter_by(log_date=today).all()
    present = sum(1 for l in t_logs if l.status == 'present')
    late = sum(1 for l in t_logs if l.status == 'late')
    absent = sum(1 for l in t_logs if l.status == 'absent')
    on_leave = LeaveRequest.query.filter(
        LeaveRequest.status == 'approved',
        LeaveRequest.start_date <= today,
        LeaveRequest.end_date >= today,
    ).count()
    no_record = total - present - late - absent - on_leave
    if no_record < 0: no_record = 0
    return jsonify({
        'labels': ['حاضر', 'غائب', 'متأخر', 'إجازة', 'لم يسجل'],
        'values': [present, absent, late, on_leave, no_record],
        'colors': ['#22c55e', '#ef4444', '#f59e0b', '#3b82f6', '#9ca3af'],
        'total': total,
    })


@admin_dashboard_bp.route('/api/dashboard/charts/heatmap')
@safe_json_response
@admin_required
def api_charts_heatmap():
    today = date.today()
    depts = Department.query.filter_by(is_active=True).order_by(Department.name_ar).all()
    days = [today - timedelta(days=i) for i in range(6, -1, -1)]
    rows = []
    for dept in depts:
        emp_ids = [e.id for e in Employee.query.filter_by(department_id=dept.id, deleted_at=None, is_active=True).all()]
        if not emp_ids:
            continue
        total_emp = len(emp_ids)
        cells = []
        for d in days:
            log_count = AttendanceLog.query.filter(
                AttendanceLog.employee_id.in_(emp_ids),
                AttendanceLog.log_date == d,
                AttendanceLog.status.in_(['present', 'late']),
            ).count()
            pct = round((log_count / total_emp) * 100) if total_emp > 0 else 0
            cells.append({'date': d.strftime('%Y-%m-%d'), 'pct': pct, 'count': log_count, 'total': total_emp})
        today_pct = cells[-1]['pct'] if cells else 0
        rows.append({
            'dept_id': dept.id,
            'dept_name': dept.name_ar,
            'dept_color': dept.color or '#e53935',
            'today_pct': today_pct,
            'cells': cells,
        })
    return jsonify({'rows': rows, 'day_labels': [DAY_NAMES[d.weekday()] for d in days]})


@admin_dashboard_bp.route('/api/dashboard/charts/punctuality')
@safe_json_response
@admin_required
def api_charts_punctuality():
    today = date.today()
    month_start = today.replace(day=1)
    employees = Employee.query.filter_by(deleted_at=None, is_active=True).all()
    ranking = []
    for emp in employees:
        total_days = (today - month_start).days + 1
        attended = AttendanceLog.query.filter(
            AttendanceLog.employee_id == emp.id,
            AttendanceLog.log_date >= month_start,
            AttendanceLog.log_date <= today,
            AttendanceLog.status.in_(['present', 'late']),
        ).count()
        late_days = AttendanceLog.query.filter(
            AttendanceLog.employee_id == emp.id,
            AttendanceLog.log_date >= month_start,
            AttendanceLog.log_date <= today,
            AttendanceLog.status == 'late',
        ).count()
        punctuality = round(((attended - late_days) / total_days) * 100) if total_days > 0 else 0
        ranking.append({
            'employee_id': emp.id,
            'employee_name': emp.full_name,
            'profile_photo': emp.profile_photo,
            'attended': attended,
            'late_days': late_days,
            'punctuality': punctuality,
        })
    ranking.sort(key=lambda x: x['punctuality'], reverse=True)
    return jsonify({'ranking': ranking[:10], 'total': len(ranking)})


@admin_dashboard_bp.route('/api/dashboard/charts/hourly')
@safe_json_response
@admin_required
def api_charts_hourly():
    today = date.today()
    logs = AttendanceLog.query.filter(
        AttendanceLog.log_date == today,
        AttendanceLog.clock_in.isnot(None),
    ).all()
    hourly = {}
    for h in range(6, 23):
        hourly[h] = 0
    for log in logs:
        h = log.clock_in.hour
        if h in hourly:
            hourly[h] += 1
    data = [{'hour': f'{h:02d}:00', 'count': hourly[h]} for h in range(6, 23)]
    peak = max(hourly.values()) if hourly else 0
    peak_hour = max(hourly, key=hourly.get) if peak > 0 else 8
    return jsonify({'data': data, 'peak_hour': f'{peak_hour:02d}:00', 'peak_count': peak})


@admin_dashboard_bp.route('/api/dashboard/records')
@safe_json_response
@admin_required
def api_dashboard_records():
    today_str = request.args.get('date', '')
    try:
        log_date = date.fromisoformat(today_str) if today_str else date.today()
    except ValueError:
        log_date = date.today()
    search = request.args.get('search', '').strip()
    dept_id = request.args.get('department_id', '').strip()
    status_filter = request.args.get('status', '').strip()
    page = int(request.args.get('page', 1))
    per_page = 10
    query = db.session.query(AttendanceLog, Employee).join(
        Employee, AttendanceLog.employee_id == Employee.id
    ).filter(AttendanceLog.log_date == log_date)
    if search:
        query = query.filter(Employee.full_name.contains(search))
    if dept_id:
        query = query.filter(Employee.department_id == int(dept_id))
    if status_filter:
        query = query.filter(AttendanceLog.status == status_filter)
    total = query.count()
    query = query.order_by(AttendanceLog.clock_in.desc().nullslast())
    logs = query.offset((page - 1) * per_page).limit(per_page).all()
    items = []
    for log, emp in logs:
        device_name = ''
        if log.device_serial:
            dev = BioTimeDevice.query.filter_by(serial_number=log.device_serial).first()
            if dev:
                device_name = dev.name or dev.serial_number
        duration = ''
        if log.clock_in and log.clock_out:
            mins = int((log.clock_out - log.clock_in).total_seconds() / 60)
            hours = mins // 60
            minutes = mins % 60
            duration = f'{hours}:{minutes:02d}'
        items.append({
            'employee_id': emp.id,
            'employee_name': emp.full_name,
            'department': emp.department,
            'department_id': emp.department_id,
            'profile_photo': emp.profile_photo,
            'clock_in': log.clock_in.strftime('%H:%M') if log.clock_in else None,
            'clock_out': log.clock_out.strftime('%H:%M') if log.clock_out else None,
            'status': log.status or 'absent',
            'device_serial': log.device_serial,
            'device_name': device_name,
            'duration': duration,
        })
    has_more = (page * per_page) < total
    return jsonify({'items': items, 'total': total, 'page': page, 'has_more': has_more, 'per_page': per_page})


@admin_dashboard_bp.route('/api/dashboard/filters')
@safe_json_response
@admin_required
def api_dashboard_filters():
    depts = Department.query.filter_by(is_active=True).order_by(Department.name_ar).all()
    return jsonify({
        'departments': [{'id': d.id, 'name_ar': d.name_ar} for d in depts],
        'statuses': [
            {'value': 'present', 'label': 'حاضر'},
            {'value': 'late', 'label': 'متأخر'},
            {'value': 'absent', 'label': 'غائب'},
        ],
    })


@admin_dashboard_bp.route('/api/dashboard/alerts')
@safe_json_response
@admin_required
def api_dashboard_alerts():
    today = date.today()
    alerts = []
    offline_devices = BioTimeDevice.query.filter_by(is_active=True).filter(
        BioTimeDevice.is_online == False
    ).all()
    for dev in offline_devices:
        offline_since = ''
        if dev.last_online_at:
            mins = int((datetime.now(UTC) - dev.last_online_at).total_seconds() / 60)
            offline_since = f'{mins} دقيقة'
        alerts.append({
            'type': 'critical',
            'icon': 'alert-circle',
            'title': f'جهاز البصمة {dev.name or dev.serial_number} غير متصل',
            'message': f'منذ {offline_since}' if offline_since else 'غير متصل',
            'link': '/admin/devices',
        })
    expiring_docs = EmployeeDocument.query.filter(
        EmployeeDocument.expiry_date.isnot(None),
        EmployeeDocument.expiry_date <= today + timedelta(days=7),
        EmployeeDocument.expiry_date > today,
    ).count()
    if expiring_docs > 0:
        alerts.append({
            'type': 'warning',
            'icon': 'file-alert',
            'title': f'{expiring_docs} مستندات تنتهي خلال 7 أيام',
            'message': 'يرجى مراجعة وتجديد المستندات',
            'link': '/admin/documents',
        })
    total = Employee.query.filter_by(deleted_at=None, is_active=True).count()
    late_count = sum(1 for l in AttendanceLog.query.filter_by(log_date=today).all() if l.status == 'late')
    if late_count > 0:
        alerts.append({
            'type': 'warning',
            'icon': 'clock-alert',
            'title': f'{late_count} موظف متأخر اليوم',
            'message': 'تجاوزوا وقت الدوام المحدد',
            'link': '/admin/attendance?status=late',
        })
    no_clock = total - len(set(l.employee_id for l in AttendanceLog.query.filter_by(log_date=today).all() if l.clock_in))
    if no_clock > 3:
        alerts.append({
            'type': 'info',
            'icon': 'user-x',
            'title': f'{no_clock} موظف لم يسجل حضورهم بعد',
            'message': 'لم يسجلوا حضورهم حتى الآن',
            'link': '/admin/attendance',
        })
    understaffed = []
    depts = Department.query.filter_by(is_active=True).all()
    for d in depts:
        if d.min_staff_required and d.min_staff_required > 0:
            emp_count = Employee.query.filter_by(department_id=d.id, deleted_at=None, is_active=True).count()
            if emp_count < d.min_staff_required:
                understaffed.append(f'{d.name_ar} ({emp_count}/{d.min_staff_required})')
    if understaffed:
        alerts.append({
            'type': 'critical',
            'icon': 'users-minus',
            'title': 'أقسام تحت الحد الأدنى للموظفين',
            'message': '، '.join(understaffed[:3]),
            'link': '/admin/departments',
        })
    return jsonify({'alerts': alerts, 'count': len(alerts)})


@admin_dashboard_bp.route('/api/dashboard/schedule')
@safe_json_response
@admin_required
def api_dashboard_schedule():
    today = date.today()
    now = datetime.now(UTC)
    current_hour = now.hour
    shifts = []
    if 6 <= current_hour < 14:
        current_shift = 'morning'
    elif 14 <= current_hour < 22:
        current_shift = 'evening'
    else:
        current_shift = 'night'
    shift_defs = [
        {'id': 'morning', 'label': 'صباحية', 'start': 6, 'end': 14, 'color': '#22c55e'},
        {'id': 'evening', 'label': 'مسائية', 'start': 14, 'end': 22, 'color': '#3b82f6'},
        {'id': 'night', 'label': 'ليلية', 'start': 22, 'end': 6, 'color': '#8b5cf6'},
    ]
    for sd in shift_defs:
        if sd['start'] < sd['end']:
            in_range = sd['start'] <= current_hour < sd['end']
        else:
            in_range = current_hour >= sd['start'] or current_hour < sd['end']
        st = ShiftType.query.filter_by(name_ar=sd['label']).first()
        scheduled_count = Employee.query.filter_by(deleted_at=None, is_active=True,
            shift_type_id=st.id if st else None
        ).count()
        if sd['start'] < sd['end']:
            clocked_in = AttendanceLog.query.filter(
                AttendanceLog.log_date == today,
                AttendanceLog.status.in_(['present', 'late']),
            ).count()
        else:
            prev_day = today - timedelta(days=1)
            clocked_in = AttendanceLog.query.filter(
                AttendanceLog.log_date == prev_day,
                AttendanceLog.clock_in.isnot(None),
                AttendanceLog.clock_out.is_(None),
            ).count()
        if sd['id'] == 'night' and current_hour < 6:
            clocked_in = AttendanceLog.query.filter(
                AttendanceLog.log_date == today,
                AttendanceLog.clock_in.isnot(None),
            ).count()
        color = sd['color']
        if in_range:
            is_sufficient = clocked_in >= max(1, scheduled_count // 2)
            bar_color = 'var(--green)' if is_sufficient else 'var(--red)'
        else:
            bar_color = 'var(--border)'
        next_shift = None
        for i, s in enumerate(shift_defs):
            if s['id'] == current_shift:
                next_idx = (i + 1) % 3
                next_shift = shift_defs[next_idx]
                break
        next_start = ''
        if next_shift:
            next_hour = next_shift['start']
            next_dt = now.replace(hour=next_hour, minute=0, second=0, microsecond=0)
            if next_hour <= current_hour:
                next_dt += timedelta(days=1)
            remaining = int((next_dt - now).total_seconds())
            if remaining > 0:
                hrs = remaining // 3600
                mins = (remaining % 3600) // 60
                next_start = f'{hrs}:{mins:02d}'
        shifts.append({
            'id': sd['id'],
            'label': sd['label'],
            'start': f'{sd["start"]:02d}:00',
            'end': f'{sd["end"]:02d}:00',
            'scheduled': scheduled_count,
            'clocked_in': clocked_in,
            'is_current': in_range,
            'bar_color': bar_color,
            'next_shift_start': next_start,
        })
    return jsonify({'shifts': shifts, 'current_shift': current_shift, 'current_hour': current_hour})


@admin_dashboard_bp.route('/api/dashboard/search')
@safe_json_response
@admin_required
def api_dashboard_search():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify({'employees': [], 'departments': []})
    employees = Employee.query.filter(
        Employee.full_name.contains(q),
        Employee.deleted_at.is_(None),
    ).limit(10).all()
    departments = Department.query.filter(
        Department.name_ar.contains(q) | Department.name_en.contains(q),
    ).limit(5).all()
    return jsonify({
        'employees': [{'id': e.id, 'full_name': e.full_name, 'department': e.department, 'type': 'employee'} for e in employees],
        'departments': [{'id': d.id, 'name_ar': d.name_ar, 'type': 'department'} for d in departments],
    })


@admin_dashboard_bp.route('/api/dashboard/notifications')
@safe_json_response
@admin_required
def api_dashboard_notifications():
    user_id = session.get('user_id')
    notifs = Notification.query.order_by(Notification.created_at.desc()).limit(5).all()
    unread = Notification.query.filter_by(is_read=False).count()
    return jsonify({
        'notifications': [{
            'id': n.id,
            'title': html.escape(n.title or ''),
            'message': html.escape(n.message or ''),
            'ntype': n.ntype,
            'created_at': n.created_at.isoformat() if n.created_at else None,
            'is_read': n.is_read,
        } for n in notifs],
        'unread_count': unread,
    })


@admin_dashboard_bp.route('/api/dashboard/stats/live')
@safe_json_response
@admin_required
def api_dashboard_live():
    today = date.today()
    total = Employee.query.filter_by(deleted_at=None, is_active=True).count()
    t_logs = AttendanceLog.query.filter_by(log_date=today).all()
    present = sum(1 for l in t_logs if l.status in ('present', 'late'))
    late = sum(1 for l in t_logs if l.status == 'late')
    absent = total - present
    on_leave = LeaveRequest.query.filter(
        LeaveRequest.status == 'approved',
        LeaveRequest.start_date <= today,
        LeaveRequest.end_date >= today,
    ).count()
    no_clockout = sum(1 for l in t_logs if l.clock_in and not l.clock_out and l.status != 'absent')
    return jsonify({
        'present': present,
        'absent': absent,
        'late': late,
        'total': total,
        'on_leave': on_leave,
        'no_clockout': no_clockout,
    })


@admin_dashboard_bp.route('/api/dashboard/map')
@safe_json_response
@admin_required
def api_dashboard_map():
    today = date.today()
    devices = BioTimeDevice.query.filter_by(is_active=True).all()
    markers = []
    for dev in devices:
        if not dev.latitude or not dev.longitude:
            continue
        emp_count = AttendanceLog.query.filter(
            AttendanceLog.log_date == today,
            AttendanceLog.device_serial == dev.serial_number,
            AttendanceLog.clock_in.isnot(None),
        ).count()
        status = 'online' if dev.is_online else 'offline'
        if dev.is_online and emp_count == 0:
            status = 'idle'
        markers.append({
            'id': dev.id,
            'name': dev.name or dev.serial_number,
            'serial': dev.serial_number,
            'lat': dev.latitude,
            'lng': dev.longitude,
            'ip_address': dev.ip_address,
            'is_online': dev.is_online,
            'last_online_at': dev.last_online_at.isoformat() if dev.last_online_at else None,
            'last_sync': dev.last_sync.isoformat() if dev.last_sync else None,
            'employee_count': emp_count,
            'status': status,
        })
    return jsonify({'markers': markers})


@admin_dashboard_bp.route('/api/dashboard/export-records')
@safe_json_response
@admin_required
def api_dashboard_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    today = date.today()
    logs = db.session.query(AttendanceLog, Employee).join(
        Employee, AttendanceLog.employee_id == Employee.id
    ).filter(AttendanceLog.log_date == today).order_by(AttendanceLog.clock_in.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = f'kashf_{today.isoformat()}'
    ws.sheet_view.rightToLeft = True
    hdrs = ['الموظف', 'القسم', 'وقت الحضور', 'وقت الانصراف', 'الحالة', 'المدة']
    hfill = PatternFill('solid', fgColor='DC2626')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = Alignment(horizontal='center')
    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill('solid', fgColor='FEF2F2')
    for ri, (log, emp) in enumerate(logs, 2):
        duration = ''
        if log.clock_in and log.clock_out:
            mins = int((log.clock_out - log.clock_in).total_seconds() / 60)
            duration = f'{mins//60}:{mins%60:02d}'
        vals = [emp.full_name, emp.department,
                log.clock_in.strftime('%H:%M') if log.clock_in else '',
                log.clock_out.strftime('%H:%M') if log.clock_out else '',
                log.status or '', duration]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border; c.alignment = Alignment(horizontal='center')
            if ri % 2 == 0: c.fill = alt_fill
    from flask import send_file
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'kashf_{today.isoformat()}.xlsx')
