import io, json, calendar
from datetime import datetime, date, timedelta, UTC
from collections import defaultdict

from flask import (Blueprint, render_template, request, session,
                   jsonify, send_file, current_app)
from models import db, Employee, ShiftType, ShiftSchedule, ShiftSwapRequest, ShiftCoverageRule, ShiftException
from utils.decorators import admin_required
from utils.helpers import coverage_status, check_conflict, safe_json
from utils.constants import MONTH_NAMES, DAY_NAMES
from services.shift_service import (check_employee_availability, validate_coverage,
                                     auto_find_substitute, resolve_conflicts_for_date,
                                     apply_leave_conflicts)
from sqlalchemy import extract
from functools import wraps

LOGGER = logging.getLogger(__name__)


def safe_api(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except Exception as e:
            LOGGER.error('API error in %s: %s', f.__name__, e)
            return jsonify({'ok': False, 'msg': str(e)}), 500
    return wrapper


admin_shifts_bp = Blueprint('admin_shifts_bp', __name__)


@admin_shifts_bp.route('/admin/shifts')
@admin_required
def admin_shifts():
    today = date.today()
    month = request.args.get('month', today.month, type=int)
    year  = request.args.get('year',  today.year,  type=int)
    dept  = request.args.get('dept', '')
    _, days_in_month = calendar.monthrange(year, month)
    cal_weeks = calendar.monthcalendar(year, month)
    month_start = date(year, month, 1)
    month_end   = date(year, month, days_in_month)
    q = ShiftSchedule.query.filter(
        ShiftSchedule.scheduled_date >= month_start,
        ShiftSchedule.scheduled_date <= month_end,
        ShiftSchedule.status == 'confirmed')
    schedules = q.all()
    day_map = {}
    for s in schedules:
        d = s.scheduled_date
        if d not in day_map: day_map[d] = {}
        st = s.shift_type_id
        if st not in day_map[d]: day_map[d][st] = []
        emp_obj = Employee.query.get(s.employee_id)
        if emp_obj and (not dept or emp_obj.department == dept):
            day_map[d][st].append({'emp': emp_obj, 'sched_id': s.id})
    shift_types  = ShiftType.query.filter_by(is_active=True).all()
    employees    = Employee.query.filter_by(role='employee', is_active=True).order_by(Employee.department).all()
    departments  = [d[0] for d in db.session.query(Employee.department).distinct().all()]
    total_assigned = len(schedules)
    pending_swaps  = ShiftSwapRequest.query.filter_by(status='pending').count()
    coverage_warnings = []
    for d_offset in range(days_in_month):
        d = date(year, month, d_offset + 1)
        for st in shift_types:
            count = len(day_map.get(d, {}).get(st.id, []))
            if count < st.min_staff:
                coverage_warnings.append({'date': d, 'shift': st.name, 'have': count, 'need': st.min_staff})
    shift_type_counts = {}
    for st in shift_types:
        cnt = 0
        for d_shifts in day_map.values():
            cnt += len(d_shifts.get(st.id, []))
        shift_type_counts[st.id] = cnt
    cal_dates = {d: date(year, month, d) for d in range(1, days_in_month + 1)}
    return render_template('admin/shifts.html',
        month=month, year=year, dept=dept, month_name=MONTH_NAMES[month-1],
        months=MONTH_NAMES, cal_weeks=cal_weeks, days_in_month=days_in_month,
        day_map=day_map, shift_types=shift_types, employees=employees,
        departments=departments, total_assigned=total_assigned,
        pending_swaps=pending_swaps, coverage_warnings=coverage_warnings[:10],
        today=today, shift_type_counts=shift_type_counts, cal_dates=cal_dates)


@admin_shifts_bp.route('/admin/shifts/types')
@admin_required
def admin_shift_types():
    types = ShiftType.query.order_by(ShiftType.start_hour).all()
    for st in types:
        st.usage_count = ShiftSchedule.query.filter_by(shift_type_id=st.id, status='confirmed').count()
    return render_template('admin/shift_types.html', types=types)


@admin_shifts_bp.route('/admin/shifts/types/add', methods=['POST'])
@admin_required
def add_shift_type():
    d = request.get_json() or {}
    if ShiftType.query.filter_by(name=d['name']).first():
        return jsonify({'ok': False, 'msg': 'اسم المناوبة موجود مسبقاً.'})
    st = ShiftType(name=d['name'],
                   start_hour=int(d['start_hour']), start_min=int(d.get('start_min', 0)),
                   end_hour=int(d['end_hour']),     end_min=int(d.get('end_min', 0)),
                   color=d.get('color', '#3b82f6'), description=d.get('description', ''),
                   min_staff=int(d.get('min_staff', 1)), max_staff=int(d.get('max_staff', 10)),
                   is_overnight=d.get('overnight', False))
    db.session.add(st); db.session.commit()
    return jsonify({'ok': True, 'msg': f'تم إضافة مناوبة "{st.name}" بنجاح.'})


@admin_shifts_bp.route('/admin/shifts/types/<int:tid>/edit', methods=['POST'])
@admin_required
def edit_shift_type(tid):
    st = ShiftType.query.get_or_404(tid)
    d  = request.get_json() or {}
    for k in ('name', 'start_hour', 'start_min', 'end_hour', 'end_min', 'color', 'description', 'min_staff', 'max_staff'):
        if k in d: setattr(st, k, d[k] if isinstance(d[k], str) else int(d[k]))
    if 'overnight' in d: st.is_overnight = bool(d['overnight'])
    db.session.commit()
    return jsonify({'ok': True, 'msg': 'تم تحديث المناوبة.'})


@admin_shifts_bp.route('/admin/shifts/types/<int:tid>/toggle', methods=['POST'])
@admin_required
def toggle_shift_type(tid):
    st = ShiftType.query.get_or_404(tid)
    st.is_active = not st.is_active
    db.session.commit()
    return jsonify({'ok': True, 'msg': f'{"فعّال" if st.is_active else "معطّل"}'})


@admin_shifts_bp.route('/admin/shifts/assign', methods=['POST'])
@admin_required
def assign_shift():
    d    = request.get_json() or {}
    emp_id  = int(d['employee_id'])
    sched_d = datetime.strptime(d['date'], '%Y-%m-%d').date()
    st_id   = int(d['shift_type_id'])
    if check_conflict(emp_id, sched_d):
        emp = Employee.query.get(emp_id)
        return jsonify({'ok': False, 'msg': f'للموظف {emp.full_name} مناوبة في نفس اليوم!'})
    existing = ShiftSchedule.query.filter_by(
        employee_id=emp_id, scheduled_date=sched_d, shift_type_id=st_id).first()
    if existing:
        existing.status = 'confirmed'
    else:
        ss = ShiftSchedule(employee_id=emp_id, shift_type_id=st_id,
                           scheduled_date=sched_d, status='confirmed',
                           notes=d.get('notes', ''), created_by=session['user_id'])
        db.session.add(ss)
    db.session.commit()
    emp = Employee.query.get(emp_id)
    st  = ShiftType.query.get(st_id)
    return jsonify({'ok': True, 'msg': f'✓ تم تعيين {emp.full_name} في مناوبة {st.name}.'})


@admin_shifts_bp.route('/admin/shifts/assign/bulk', methods=['POST'])
@admin_required
def bulk_assign_shift():
    d        = request.get_json() or {}
    st_id    = int(d['shift_type_id'])
    emp_ids  = [int(i) for i in d.get('employee_ids', [])]
    dates    = [datetime.strptime(x, '%Y-%m-%d').date() for x in d.get('dates', [])]
    skip_fri = d.get('skip_friday', True)
    skip_sat = d.get('skip_saturday', True)
    if not emp_ids or not dates:
        return jsonify({'ok': False, 'msg': 'يرجى تحديد الموظفين والتواريخ.'})
    count = 0; skipped = 0
    for emp_id in emp_ids:
        for sched_date in dates:
            if skip_fri and sched_date.weekday() == 4: continue
            if skip_sat and sched_date.weekday() == 5: continue
            if check_conflict(emp_id, sched_date): skipped += 1; continue
            existing = ShiftSchedule.query.filter_by(
                employee_id=emp_id, scheduled_date=sched_date, shift_type_id=st_id).first()
            if not existing:
                db.session.add(ShiftSchedule(
                    employee_id=emp_id, shift_type_id=st_id,
                    scheduled_date=sched_date, status='confirmed',
                    created_by=session['user_id']))
                count += 1
    db.session.commit()
    msg = f'✓ تم تعيين {count} مناوبة.'
    if skipped: msg += f' ({skipped} تجاوزت بسبب تعارض أو عطلة)'
    return jsonify({'ok': True, 'msg': msg, 'count': count})


@admin_shifts_bp.route('/admin/shifts/copy-week', methods=['POST'])
@admin_required
def copy_week():
    d = request.get_json() or {}
    src_start = datetime.strptime(d['source_start'], '%Y-%m-%d').date()
    tgt_start = datetime.strptime(d['target_start'], '%Y-%m-%d').date()
    src_end   = src_start + timedelta(days=6)
    src_schedules = ShiftSchedule.query.filter(
        ShiftSchedule.scheduled_date >= src_start,
        ShiftSchedule.scheduled_date <= src_end,
        ShiftSchedule.status == 'confirmed').all()
    count = 0; skipped = 0
    for ss in src_schedules:
        delta    = ss.scheduled_date - src_start
        new_date = tgt_start + delta
        if check_conflict(ss.employee_id, new_date): skipped += 1; continue
        existing = ShiftSchedule.query.filter_by(
            employee_id=ss.employee_id, scheduled_date=new_date,
            shift_type_id=ss.shift_type_id).first()
        if not existing:
            db.session.add(ShiftSchedule(
                employee_id=ss.employee_id, shift_type_id=ss.shift_type_id,
                scheduled_date=new_date, status='confirmed',
                notes='نسخة من أسبوع سابق', created_by=session['user_id']))
            count += 1
    db.session.commit()
    return jsonify({'ok': True, 'msg': f'✓ تم نسخ {count} مناوبة للأسبوع الجديد. (تجاوز: {skipped})', 'count': count})


@admin_shifts_bp.route('/admin/shifts/auto-rotate', methods=['POST'])
@admin_required
def auto_rotate():
    d        = request.get_json() or {}
    dept     = d.get('department')
    st_ids   = [int(x) for x in d.get('shift_type_ids', [])]
    start_d  = datetime.strptime(d['start_date'], '%Y-%m-%d').date()
    end_d    = datetime.strptime(d['end_date'],   '%Y-%m-%d').date()
    skip_fri = d.get('skip_friday', True)
    skip_sat = d.get('skip_saturday', True)
    if not dept or not st_ids:
        return jsonify({'ok': False, 'msg': 'يرجى تحديد القسم وأنواع المناوبات.'})
    employees = Employee.query.filter_by(department=dept, role='employee', is_active=True).all()
    if not employees:
        return jsonify({'ok': False, 'msg': 'لا يوجد موظفون في هذا القسم.'})
    shift_types = [ShiftType.query.get(sid) for sid in st_ids if ShiftType.query.get(sid)]
    if not shift_types:
        return jsonify({'ok': False, 'msg': 'أنواع المناوبات غير موجودة.'})
    count = 0; skipped = 0
    n_shifts = len(shift_types)
    curr = start_d
    shift_idx = 0
    while curr <= end_d:
        if (skip_fri and curr.weekday() == 4) or (skip_sat and curr.weekday() == 5):
            curr += timedelta(days=1); continue
        for i, emp in enumerate(employees):
            st = shift_types[(shift_idx + i) % n_shifts]
            if check_conflict(emp.id, curr): skipped += 1; continue
            existing = ShiftSchedule.query.filter_by(
                employee_id=emp.id, scheduled_date=curr,
                shift_type_id=st.id, status='confirmed').first()
            if not existing:
                db.session.add(ShiftSchedule(
                    employee_id=emp.id, shift_type_id=st.id,
                    scheduled_date=curr, status='confirmed',
                    notes='تدوير تلقائي', created_by=session['user_id']))
                count += 1
        shift_idx = (shift_idx + 1) % n_shifts
        curr += timedelta(days=1)
    db.session.commit()
    return jsonify({'ok': True, 'msg': f'✓ تم إنشاء {count} مناوبة بالتدوير التلقائي.', 'count': count})


@admin_shifts_bp.route('/admin/shifts/<int:sid>/cancel', methods=['POST'])
@admin_required
def cancel_shift(sid):
    ss = ShiftSchedule.query.get_or_404(sid)
    ss.status = 'cancelled'
    db.session.commit()
    return jsonify({'ok': True, 'msg': 'تم إلغاء المناوبة.'})


@admin_shifts_bp.route('/api/shifts/day/<date_str>')
@admin_required
@safe_api
def api_day_shifts(date_str):
    d = datetime.strptime(date_str, '%Y-%m-%d').date()
    shift_types = ShiftType.query.filter_by(is_active=True).order_by(ShiftType.start_hour).all()
    result = []
    for st in shift_types:
        schedules = ShiftSchedule.query.filter_by(
            shift_type_id=st.id, scheduled_date=d, status='confirmed').all()
        emps = []
        for ss in schedules:
            emp = Employee.query.get(ss.employee_id)
            if emp: emps.append({'id': emp.id, 'name': emp.full_name,
                                  'dept': emp.department, 'sched_id': ss.id})
        result.append({
            'shift_id': st.id, 'shift_name': st.name,
            'time_range': st.time_range, 'color': st.color,
            'min_staff': st.min_staff, 'employees': emps,
            'coverage': 'ok' if len(emps) >= st.min_staff else ('low' if emps else 'empty')})
    all_employees = [{'id': e.id, 'name': e.full_name, 'dept': e.department}
                     for e in Employee.query.filter_by(role='employee', is_active=True)
                                       .order_by(Employee.department).all()]
    return jsonify({'date': date_str, 'shifts': result, 'all_employees': all_employees})


@admin_shifts_bp.route('/admin/shifts/clear-day', methods=['POST'])
@admin_required
def clear_day():
    d = request.get_json() or {}
    date_str = d.get('date', '')
    if date_str:
        sched_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        ShiftSchedule.query.filter_by(scheduled_date=sched_date, status='confirmed').delete()
        db.session.commit()
    return jsonify({'ok': True, 'msg': '✓ تم تعيين اليوم كراحة إلزامية وإلغاء جميع المناوبات.'})


@admin_shifts_bp.route('/admin/shifts/swaps')
@admin_required
def admin_shift_swaps():
    status  = request.args.get('status', 'pending')
    swaps   = ShiftSwapRequest.query.filter_by(status=status)\
                .order_by(ShiftSwapRequest.created_at.desc()).all()
    counts  = {
        'pending':        ShiftSwapRequest.query.filter_by(status='pending').count(),
        'target_approved':ShiftSwapRequest.query.filter_by(status='target_approved').count(),
        'approved':       ShiftSwapRequest.query.filter_by(status='approved').count(),
        'rejected':       ShiftSwapRequest.query.filter_by(status='rejected').count(),
    }
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.department, Employee.full_name).all()
    departments = sorted(set(e.department for e in employees))
    return render_template('admin/shift_swaps.html', swaps=swaps,
                           status=status, counts=counts,
                           employees=employees, departments=departments)


@admin_shifts_bp.route('/admin/shifts/swaps/<int:swap_id>/action', methods=['POST'])
@admin_required
def admin_swap_action(swap_id):
    swap  = ShiftSwapRequest.query.get_or_404(swap_id)
    d     = request.get_json() or {}
    action= d.get('action')
    if action == 'approve':
        if swap.req_sched and swap.tgt_sched:
            req_s = ShiftSchedule.query.get(swap.req_sched_id)
            tgt_s = ShiftSchedule.query.get(swap.tgt_sched_id)
            if req_s and tgt_s:
                if not req_s.original_employee_id:
                    req_s.original_employee_id = req_s.employee_id
                if not tgt_s.original_employee_id:
                    tgt_s.original_employee_id = tgt_s.employee_id
                req_s.employee_id, tgt_s.employee_id = tgt_s.employee_id, req_s.employee_id
                req_s.notes = 'تبديل مناوبة معتمد'
                tgt_s.notes = 'تبديل مناوبة معتمد'
        swap.status      = 'approved'
        swap.admin_notes = d.get('notes', '')
        swap.updated_at  = datetime.now(UTC)
        db.session.commit()
        return jsonify({'ok': True, 'msg': '✓ تم اعتماد التبديل وتحديث الجدول.'})
    elif action == 'reject':
        swap.status      = 'rejected'
        swap.admin_notes = d.get('notes', 'مرفوض من قِبل الإدارة')
        swap.updated_at  = datetime.now(UTC)
        db.session.commit()
        return jsonify({'ok': True, 'msg': 'تم رفض طلب التبديل.'})
    return jsonify({'ok': False, 'msg': 'إجراء غير معروف.'})


@admin_shifts_bp.route('/api/shifts/availability/<int:emp_id>/<date_str>')
@admin_required
@safe_api
def api_employee_availability(emp_id, date_str):
    sched_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    conflicts = check_employee_availability(emp_id, sched_date)
    return jsonify({
        'available': len(conflicts) == 0,
        'conflicts': [{'type': c[0], 'msg': c[1]} for c in conflicts]
    })


@admin_shifts_bp.route('/admin/shifts/substitute', methods=['POST'])
@admin_required
def assign_substitute():
    d = request.get_json() or {}
    absent_emp_id = int(d['absent_employee_id'])
    substitute_emp_id = int(d['substitute_employee_id'])
    sched_date = datetime.strptime(d['date'], '%Y-%m-%d').date()
    shift_type_id = int(d['shift_type_id'])

    original_schedule = ShiftSchedule.query.filter_by(
        employee_id=absent_emp_id, scheduled_date=sched_date, status='confirmed'
    ).first()

    if not original_schedule:
        return jsonify({'ok': False, 'msg': 'الموظف الأصلي ليس لديه مناوبة في هذا اليوم.'})

    target_conflicts = check_employee_availability(substitute_emp_id, sched_date)
    if target_conflicts:
        msg = '; '.join(c[1] for c in target_conflicts)
        return jsonify({'ok': False, 'msg': f'الموظف البديل غير متاح: {msg}'})

    original_schedule.original_employee_id = absent_emp_id
    original_schedule.employee_id = substitute_emp_id
    original_schedule.status = 'substituted'
    original_schedule.notes = d.get('notes', 'تعيين بديل')
    original_schedule.substituted_by = session['user_id']
    original_schedule.substituted_at = datetime.now(UTC)

    db.session.add(ShiftException(
        employee_id=absent_emp_id,
        shift_schedule_id=original_schedule.id,
        exception_date=sched_date,
        exception_type='admin',
        reason=d.get('notes', 'تعيين بديل إداري')
    ))
    db.session.commit()

    absent = Employee.query.get(absent_emp_id)
    substitute = Employee.query.get(substitute_emp_id)
    return jsonify({
        'ok': True,
        'msg': f'✓ تم تعيين {substitute.full_name} بديلاً عن {absent.full_name} في مناوبة {original_schedule.shift_type.name}.'
    })


@admin_shifts_bp.route('/api/shifts/substitute-candidates/<int:sched_id>')
@admin_required
@safe_api
def api_substitute_candidates(sched_id):
    dept = request.args.get('dept')
    candidates = auto_find_substitute(sched_id, department=dept)
    return jsonify({'candidates': candidates})


@admin_shifts_bp.route('/api/shifts/resolve-conflicts/<date_str>', methods=['POST'])
@admin_required
@safe_api
def api_resolve_conflicts(date_str):
    sched_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    updated = resolve_conflicts_for_date(sched_date)
    return jsonify({'ok': True, 'updated': updated})


@admin_shifts_bp.route('/admin/shifts/export')
@admin_required
def export_shifts_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    today = date.today()
    month = request.args.get('month', today.month, type=int)
    year  = request.args.get('year',  today.year,  type=int)
    _, days_in_month = calendar.monthrange(year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = f"مناوبات {MONTH_NAMES[month-1]} {year}"
    ws.sheet_view.rightToLeft = True
    ws.merge_cells(f'A1:{get_column_letter(days_in_month + 1)}1')
    t = ws['A1']
    t.value = f"جدول المناوبات — {MONTH_NAMES[month-1]} {year} — SMARTLOG"
    t.font  = Font(bold=True, size=13, color='FFFFFF')
    t.fill  = PatternFill("solid", fgColor='991B1B')
    t.alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=1, value='الموظف').font = Font(bold=True)
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        is_fri = d.weekday() == 4
        is_sat = d.weekday() == 5
        c = ws.cell(row=2, column=day+1, value=f"{day}\n{['الأحد','الاثنين','الثلاثاء','الأربعاء','الخميس','الجمعة','السبت'][d.weekday()][:3]}")
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.font  = Font(bold=True, color='FFFFFF' if (is_fri or is_sat) else '000000')
        if is_fri or is_sat: c.fill = PatternFill("solid", fgColor='475569')
    employees = Employee.query.filter_by(role='employee', is_active=True)\
                    .order_by(Employee.department, Employee.full_name).all()
    shift_colors = {}
    for st in ShiftType.query.all():
        hex_c = st.color.lstrip('#')
        shift_colors[st.id] = (hex_c, st.name[:3])
    for ri, emp in enumerate(employees, 3):
        ws.cell(row=ri, column=1, value=f"{emp.full_name} ({emp.department})")
        ws.cell(row=ri, column=1).alignment = Alignment(horizontal='right')
        for day in range(1, days_in_month + 1):
            d = date(year, month, day)
            ss = ShiftSchedule.query.filter_by(
                employee_id=emp.id, scheduled_date=d, status='confirmed').first()
            if ss:
                hex_c, short = shift_colors.get(ss.shift_type_id, ('3b82f6', '؟'))
                c = ws.cell(row=ri, column=day+1, value=short)
                c.fill = PatternFill("solid", fgColor=hex_c.upper())
                c.font = Font(color='FFFFFF', bold=True, size=9)
                c.alignment = Alignment(horizontal='center')
    for col in range(1, days_in_month + 2):
        ws.column_dimensions[get_column_letter(col)].width = 5 if col > 1 else 28
    for row in range(1, len(employees) + 3):
        ws.row_dimensions[row].height = 22
    out = io.BytesIO(); wb.save(out); out.seek(0)
    fname = f'مناوبات_{MONTH_NAMES[month-1]}_{year}.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@admin_shifts_bp.route('/admin/shifts/coverage')
@admin_required
def shift_coverage():
    today = date.today()
    month = request.args.get('month', today.month, type=int)
    year  = request.args.get('year',  today.year,  type=int)
    _, days = calendar.monthrange(year, month)
    shift_types = ShiftType.query.filter_by(is_active=True).all()
    data = []
    for st in shift_types:
        row = {'type': st, 'days': []}
        for day in range(1, days + 1):
            d     = date(year, month, day)
            count = ShiftSchedule.query.filter_by(
                shift_type_id=st.id, scheduled_date=d, status='confirmed').count()
            status, _ = coverage_status(count, st.min_staff)
            conflict_count = ShiftSchedule.query.filter_by(
                shift_type_id=st.id, scheduled_date=d, status='confirmed'
            ).filter(
                ShiftSchedule.conflict_status != 'ok'
            ).count()
            row['days'].append({
                'day': day, 'count': count, 'status': status, 'dow': d.weekday(),
                'conflict_count': conflict_count,
            })
        data.append(row)
    employees_q = Employee.query.filter_by(is_active=True, role='employee').order_by(Employee.full_name).all()
    shift_types_q = ShiftType.query.filter_by(is_active=True).all()
    return render_template('admin/shift_coverage.html',
        data=data, month=month, year=year,
        month_name=MONTH_NAMES[month-1], months=MONTH_NAMES, days=days, today=today,
        employees=employees_q, shift_types=shift_types_q)
