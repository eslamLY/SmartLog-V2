import json, io, calendar
from datetime import datetime, date, timedelta, UTC
from collections import defaultdict

from flask import Blueprint, request, jsonify, render_template, session, send_file
from sqlalchemy import func, extract

from models import db
from models.employee import Employee
from models.attendance import AttendanceLog
from models.attendance_report import ReportDataService, ReportCorrection, ScheduledReport
from models.misc import LeaveRequest
from models.shifts import ShiftType
from models.department import Department
from services.payroll_service import PayrollService
from utils.decorators import admin_required

admin_reports_bp = Blueprint('admin_reports', __name__, url_prefix='/admin/reports')

MONTH_NAMES = ['','يناير','فبراير','مارس','أبريل','مايو','يونيو','يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر']


@admin_reports_bp.route('')
@admin_required
def reports_page():
    today = date.today()
    departments = Department.query.filter_by(is_active=True).order_by(Department.name_ar).all()
    shifts = ShiftType.query.filter_by(is_active=True).all()
    month = request.args.get('month', today.month, type=int)
    year = request.args.get('year', today.year, type=int)
    months = [{'value': i, 'label': f'{i:02d}', 'selected': i == month} for i in range(1, 13)]
    years = list(range(today.year - 2, today.year + 3))
    return render_template('admin/reports.html',
        departments=departments,
        shifts=shifts,
        months=months,
        years=years,
        now=today,
    )


@admin_reports_bp.route('/api/data')
@admin_required
def api_report_data():
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    dept_id = request.args.get('department_id', type=int)
    employee_id = request.args.get('employee_id', type=int)
    emp_type = request.args.get('employment_type', '').strip() or None
    shift_id = request.args.get('shift_id', type=int)
    show_present = request.args.get('show_present', '1') == '1'
    show_absent = request.args.get('show_absent', '1') == '1'
    show_late = request.args.get('show_late', '1') == '1'
    show_leave_only = request.args.get('show_leave_only') == '1'
    show_deductions_only = request.args.get('show_deductions_only') == '1'
    show_excess_absence = request.args.get('show_excess_absence') == '1'
    result = ReportDataService.calculate_report(year, month, dept_id, employee_id, emp_type, shift_id)
    rows = result['rows']
    summary = result['summary']
    filtered = []
    for r in rows:
        if not show_present and r['attendance_pct'] >= 95:
            continue
        if not show_absent and r['absent'] > 0:
            continue
        if not show_late and r['late_count'] > 0:
            continue
        if show_leave_only and r['leave_days'] == 0:
            continue
        if show_deductions_only and r['total_deduction'] <= 0:
            continue
        if show_excess_absence and r['absent'] < 3:
            continue
        filtered.append(r)
    serialized = []
    for r in filtered:
        serialized.append({
            'id': r['emp_id'],
            'emp_name': r['emp_name'],
            'emp_code': r['emp_code'],
            'department': r['department'],
            'department_id': r['department_id'],
            'employment_type': r['employment_type'],
            'present': r['present'],
            'late_count': r['late_count'],
            'absent': r['absent'],
            'late_minutes': r['late_minutes'],
            'expected_days': r['expected_days'],
            'total_work_hours': r['total_work_hours'],
            'total_work_hours_str': r['total_work_hours_str'],
            'leave_days': r['leave_days'],
            'overtime_minutes': r['overtime_minutes'],
            'overtime_pay': r['overtime_pay'],
            'bonus': r['bonus'],
            'late_deduction': r['late_deduction'],
            'absence_deduction': r['absence_deduction'],
            'total_deduction': r['total_deduction'],
            'base_salary': r['base_salary'],
            'allowances': r['allowances'],
            'net_salary': r['net_salary'],
            'attendance_pct': r['attendance_pct'],
            'status_label': r['status_label'],
            'overall_status': r['overall_status'],
            'profile_photo': r['profile_photo'],
            'day_details': r['day_details'],
        })
    return jsonify({
        'rows': serialized,
        'summary': summary,
        'year': year,
        'month': month,
        'month_name': MONTH_NAMES[month] if 1 <= month <= 12 else '',
        'total_filtered': len(serialized),
    })


@admin_reports_bp.route('/api/filters')
@admin_required
def api_report_filters():
    departments = Department.query.filter_by(is_active=True).order_by(Department.name_ar).all()
    shifts = ShiftType.query.filter_by(is_active=True).all()
    return jsonify({
        'departments': [{'id': d.id, 'name_ar': d.name_ar} for d in departments],
        'shifts': [{'id': s.id, 'name_ar': s.name_ar} for s in shifts],
        'employment_types': [
            {'value': 'full_time', 'label': 'دوام كامل'},
            {'value': 'part_time', 'label': 'دوام جزئي'},
            {'value': 'contract', 'label': 'عقد'},
            {'value': 'temporary', 'label': 'مؤقت'},
        ],
        'years': list(range(date.today().year - 2, date.today().year + 3)),
        'months': [{'num': i, 'name': MONTH_NAMES[i]} for i in range(1, 13)],
    })


@admin_reports_bp.route('/api/employees')
@admin_required
def api_report_employees():
    q = request.args.get('q', '').strip()
    query = Employee.query.filter_by(is_active=True, deleted_at=None)
    if q:
        query = query.filter(Employee.full_name.contains(q) | Employee.username.contains(q))
    emps = query.order_by(Employee.full_name).limit(20).all()
    return jsonify({'employees': [{'id': e.id, 'full_name': e.full_name, 'username': e.username, 'department': e.department} for e in emps]})


@admin_reports_bp.route('/api/comparison')
@admin_required
def api_report_comparison():
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    emp_ids = request.args.getlist('employee_ids[]')
    if not emp_ids:
        return jsonify({'error': 'اختر موظفين للمقارنة'}), 400
    emp_ids = [int(x) for x in emp_ids]
    results = []
    for eid in emp_ids:
        result = ReportDataService.calculate_report(year, month, employee_id=eid)
        for r in result['rows']:
            results.append(r)
    return jsonify({'rows': results, 'month_name': MONTH_NAMES[month] if 1 <= month <= 12 else ''})


@admin_reports_bp.route('/api/charts')
@admin_required
def api_report_charts():
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    dept_id = request.args.get('department_id', type=int)
    month_days = calendar.monthrange(year, month)[1]
    qry = Employee.query.filter_by(is_active=True, deleted_at=None)
    if dept_id:
        qry = qry.filter_by(department_id=dept_id)
    employees = qry.all()
    emp_ids = [e.id for e in employees]
    all_logs = AttendanceLog.query.filter(
        AttendanceLog.employee_id.in_(emp_ids) if emp_ids else False,
        extract('month', AttendanceLog.log_date) == month,
        extract('year', AttendanceLog.log_date) == year,
    ).all() if emp_ids else []
    logs_by_day = defaultdict(list)
    for l in all_logs:
        logs_by_day[l.log_date.day].append(l)
    daily_trend = []
    for day_num in range(1, month_days + 1):
        d = date(year, month, day_num)
        day_logs = logs_by_day.get(day_num, [])
        present = sum(1 for l in day_logs if l.status in ('present', 'late'))
        late = sum(1 for l in day_logs if l.status == 'late')
        absent = len(employees) - present
        daily_trend.append({
            'day': day_num,
            'present': present,
            'absent': absent,
            'late': late,
        })
    depts = Department.query.filter_by(is_active=True).all()
    dept_comparison = []
    for dept in depts:
        dept_emps = [e.id for e in Employee.query.filter_by(department_id=dept.id, is_active=True, deleted_at=None).all()]
        if not dept_emps:
            continue
        dept_logs = AttendanceLog.query.filter(
            AttendanceLog.employee_id.in_(dept_emps),
            extract('month', AttendanceLog.log_date) == month,
            extract('year', AttendanceLog.log_date) == year,
        ).all()
        total_possible = len(dept_emps) * month_days
        total_present = sum(1 for l in dept_logs if l.status in ('present', 'late'))
        pct = round((total_present / total_possible) * 100, 1) if total_possible > 0 else 0
        dept_comparison.append({'name': dept.name_ar, 'pct': pct, 'color': dept.color or '#e53935'})
    dept_comparison.sort(key=lambda x: x['pct'], reverse=True)
    late_dist = []
    for emp in employees:
        emp_logs = AttendanceLog.query.filter(
            AttendanceLog.employee_id == emp.id,
            extract('month', AttendanceLog.log_date) == month,
            extract('year', AttendanceLog.log_date) == year,
        ).all()
        total_late = sum(l.late_minutes or 0 for l in emp_logs)
        if total_late > 0:
            late_dist.append({'name': emp.full_name, 'late_minutes': total_late})
    late_dist.sort(key=lambda x: x['late_minutes'], reverse=True)
    heatmap = []
    for emp in employees[:30]:
        emp_logs = AttendanceLog.query.filter(
            AttendanceLog.employee_id == emp.id,
            extract('month', AttendanceLog.log_date) == month,
            extract('year', AttendanceLog.log_date) == year,
        ).all()
        logs_by_day_emp = {l.log_date.day: l for l in emp_logs}
        days = []
        for day_num in range(1, month_days + 1):
            l = logs_by_day_emp.get(day_num)
            if l:
                st = l.status
            else:
                d = date(year, month, day_num)
                st = 'absent'
                if d.weekday() >= 5:
                    st = 'weekend'
            days.append(st)
        heatmap.append({'name': emp.full_name, 'days': days})
    ded_total_late = sum(r['late_deduction'] for r in ReportDataService.calculate_report(year, month, dept_id)['rows'])
    ded_total_absence = sum(r['absence_deduction'] for r in ReportDataService.calculate_report(year, month, dept_id)['rows'])
    return jsonify({
        'daily_trend': daily_trend,
        'dept_comparison': dept_comparison,
        'late_distribution': late_dist[:20],
        'heatmap': heatmap,
        'deductions_breakdown': {
            'labels': ['خصم التأخير', 'خصم الغياب', 'خصومات أخرى'],
            'values': [round(ded_total_late, 2), round(ded_total_absence, 2), 0],
            'colors': ['#f59e0b', '#ef4444', '#9ca3af'],
        },
    })


@admin_reports_bp.route('/api/corrections', methods=['GET'])
@admin_required
def api_list_corrections():
    corrections = ReportCorrection.query.order_by(ReportCorrection.created_at.desc()).limit(50).all()
    return jsonify({'corrections': [c.to_dict() for c in corrections]})


@admin_reports_bp.route('/api/corrections/create', methods=['POST'])
@admin_required
def api_create_correction():
    data = request.get_json() or {}
    employee_id = data.get('employee_id')
    log_date_str = data.get('log_date')
    corr_type = data.get('correction_type', '')
    reason = data.get('reason', '').strip()
    if not employee_id or not log_date_str or not reason:
        return jsonify({'error': 'الموظف والتاريخ والسبب مطلوبون'}), 400
    try:
        log_date = date.fromisoformat(log_date_str)
    except (ValueError, TypeError):
        return jsonify({'error': 'تاريخ غير صحيح'}), 400
    original_log = AttendanceLog.query.filter_by(employee_id=int(employee_id), log_date=log_date).first()
    original_value = None
    if original_log:
        if corr_type == 'clock_in':
            original_value = original_log.clock_in.strftime('%H:%M') if original_log.clock_in else None
        elif corr_type == 'clock_out':
            original_value = original_log.clock_out.strftime('%H:%M') if original_log.clock_out else None
    c = ReportCorrection(
        employee_id=int(employee_id),
        log_date=log_date,
        correction_type=corr_type,
        original_value=original_value,
        corrected_value=data.get('corrected_value'),
        reason=reason,
        created_by=session.get('user_id'),
    )
    db.session.add(c)
    db.session.commit()
    return jsonify({'success': True, 'correction': c.to_dict()})


@admin_reports_bp.route('/api/corrections/<int:c_id>/review', methods=['POST'])
@admin_required
def api_review_correction(c_id):
    c = ReportCorrection.query.get_or_404(c_id)
    data = request.get_json() or {}
    status = data.get('status', 'approved')
    notes = data.get('notes', '')
    c.status = status
    c.reviewed_by = session.get('user_id')
    c.reviewed_at = datetime.now(UTC)
    c.review_notes = notes
    if status == 'approved' and c.correction_type in ('clock_in', 'clock_out'):
        log = AttendanceLog.query.filter_by(employee_id=c.employee_id, log_date=c.log_date).first()
        if log and c.corrected_value:
            from datetime import datetime as dt
            try:
                parts = c.corrected_value.split(':')
                h, m = int(parts[0]), int(parts[1])
                if c.correction_type == 'clock_in':
                    log.clock_in = log.log_date.isoformat() + f' {h:02d}:{m:02d}:00'
                    log.clock_in = dt.strptime(f'{log.log_date.isoformat()} {h:02d}:{m:02d}:00', '%Y-%m-%d %H:%M:%S')
                elif c.correction_type == 'clock_out':
                    log.clock_out = dt.strptime(f'{log.log_date.isoformat()} {h:02d}:{m:02d}:00', '%Y-%m-%d %H:%M:%S')
            except (ValueError, IndexError):
                pass
    db.session.commit()
    return jsonify({'success': True, 'correction': c.to_dict()})


@admin_reports_bp.route('/api/bonus/<int:employee_id>', methods=['POST'])
@admin_required
def api_set_bonus(employee_id):
    data = request.get_json() or {}
    amount = float(data.get('amount', 0))
    return jsonify({'success': True, 'bonus': amount})


@admin_reports_bp.route('/api/scheduled', methods=['GET'])
@admin_required
def api_list_scheduled():
    reports = ScheduledReport.query.order_by(ScheduledReport.created_at.desc()).all()
    return jsonify({'reports': [r.to_dict() for r in reports]})


@admin_reports_bp.route('/api/scheduled/create', methods=['POST'])
@admin_required
def api_create_scheduled():
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'الاسم مطلوب'}), 400
    sr = ScheduledReport(
        name=name,
        frequency=data.get('frequency', 'monthly'),
        day_of_month=int(data['day_of_month']) if data.get('day_of_month') else None,
        day_of_week=int(data['day_of_week']) if data.get('day_of_week') else None,
        time_hour=int(data.get('time_hour', 8)),
        time_minute=int(data.get('time_minute', 0)),
        format_type=data.get('format_type', 'pdf'),
        recipients_json=json.dumps(data.get('recipients', []), ensure_ascii=False),
        department_ids=json.dumps(data.get('department_ids', []), ensure_ascii=False),
        created_by=session.get('user_id'),
    )
    db.session.add(sr)
    db.session.commit()
    return jsonify({'success': True, 'report': sr.to_dict()})


@admin_reports_bp.route('/api/scheduled/<int:sr_id>/toggle', methods=['POST'])
@admin_required
def api_toggle_scheduled(sr_id):
    sr = ScheduledReport.query.get_or_404(sr_id)
    sr.is_active = not sr.is_active
    db.session.commit()
    return jsonify({'success': True, 'is_active': sr.is_active})


@admin_reports_bp.route('/api/scheduled/<int:sr_id>/delete', methods=['POST'])
@admin_required
def api_delete_scheduled(sr_id):
    sr = ScheduledReport.query.get_or_404(sr_id)
    db.session.delete(sr)
    db.session.commit()
    return jsonify({'success': True})


@admin_reports_bp.route('/api/scheduled/<int:sr_id>/run-now', methods=['POST'])
@admin_required
def api_run_scheduled(sr_id):
    sr = ScheduledReport.query.get_or_404(sr_id)
    sr.last_run_at = datetime.now(UTC)
    sr.last_status = 'running'
    db.session.commit()
    try:
        from services.report_service import ReportGenerationService
        today = date.today()
        year = today.year
        month = today.month
        result = ReportDataService.calculate_report(year, month)
        dept_ids = json.loads(sr.department_ids) if sr.department_ids else []
        if dept_ids:
            result = ReportDataService.calculate_report(year, month, dept_id=dept_ids[0])
        sr.last_status = 'success'
    except Exception as e:
        sr.last_status = f'failed: {str(e)}'
    sr.last_run_at = datetime.now(UTC)
    db.session.commit()
    return jsonify({'success': True, 'status': sr.last_status})


@admin_reports_bp.route('/api/export/pdf')
@admin_required
def api_export_pdf():
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    dept_id = request.args.get('department_id', type=int)
    emp_id = request.args.get('employee_id', type=int)
    result = ReportDataService.calculate_report(year, month, dept_id, emp_id)
    from services.report_service import ReportGenerationService
    pdf_bytes = ReportGenerationService.generate_pdf_report(result, year, month)
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'taqrir_{year}_{month:02d}.pdf',
    )


@admin_reports_bp.route('/api/export/excel')
@admin_required
def api_export_excel():
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    dept_id = request.args.get('department_id', type=int)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    result = ReportDataService.calculate_report(year, month, dept_id)
    wb = Workbook()
    ws = wb.active
    ws.title = f'{month:02d}-{year}'
    ws.sheet_view.rightToLeft = True
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = f'تقرير الحضور والانصراف — {month:02d}-{year} — بنك دم طبرق'
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor='991B1B')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    hdrs = ['#', 'الموظف', 'القسم', 'حضور', 'تأخير', 'غياب', 'دقائق تأخير', 'أيام متوقعة', 'خصم التأخير', 'خصم الغياب', 'إجمالي الخصم', 'صافي الراتب']
    hfill = PatternFill('solid', fgColor='DC2626')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill('solid', fgColor='FEF2F2')
    for ri, row in enumerate(result['rows'], 3):
        vals = [ri - 2, row['emp_name'], row['department'], row['present'], row['late_count'], row['absent'], row['late_minutes'], row['expected_days'], row['late_deduction'], row['absence_deduction'], row['total_deduction'], row['net_salary']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if ri % 2 == 0:
                c.fill = alt_fill
    for ci in range(1, len(hdrs) + 1):
        ws.column_dimensions[chr(64 + ci) if ci <= 26 else 'A'].width = 18
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'taqrir_{year}_{month:02d}.xlsx')


@admin_reports_bp.route('/api/mark-anomaly', methods=['POST'])
@admin_required
def api_mark_anomaly():
    data = request.get_json() or {}
    employee_id = data.get('employee_id')
    log_date_str = data.get('log_date')
    status = data.get('status', 'reviewed')
    if not employee_id or not log_date_str:
        return jsonify({'error': 'مطلوب'}), 400
    try:
        from models.anomaly import AttendanceAnomaly
        log_date = date.fromisoformat(log_date_str)
        anomaly = AttendanceAnomaly.query.filter_by(employee_id=int(employee_id), log_date=log_date).first()
        if not anomaly:
            anomaly = AttendanceAnomaly(employee_id=int(employee_id), log_date=log_date, anomaly_type='manual_review', severity='info', description='مراجعة يدوية')
            db.session.add(anomaly)
        anomaly.status = status
        anomaly.reviewed_by = session.get('user_id')
        anomaly.reviewed_at = datetime.now(UTC)
        db.session.commit()
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    return jsonify({'success': True})


@admin_reports_bp.route('/api/export-whatsapp')
@admin_required
def api_export_whatsapp():
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int) or date.today().month
    result = ReportDataService.calculate_report(year, month)
    s = result['summary']
    msg = (
        f'📊 تقرير الحضور — {month:02d}-{year}\n'
        f'إجمالي الموظفين: {s["total_employees"]}\n'
        f'الحاضرون: {s["total_present"]} | الغائبون: {s["total_absent"]}\n'
        f'معدل الحضور: {s["overall_pct"]}%\n'
        f'إجمالي الخصومات: {s["total_deductions"]} د.ل'
    )
    return jsonify({'message': msg})
